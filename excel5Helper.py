'''
Author: Uercal
Date: 2021-06-15 14:39:35
LastEditTime: 2021-07-14 09:19:48
Description:  Excel处理方法类
'''
import os
from sys import argv, path

from PyQt5.QtCore import right
import helper
#
import win32com.client as win32
import xlwt
import json
import openpyxl
import random
# 多线程
from threading import Thread
# 多进程
import multiprocessing as mp

# 根据给定索引集 进行websheet读取写入numbers


def readWbFromIndex(ws, indexes, set_range):
    #
    global configuration
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    isRange = configuration['isRange']
    #
    numbers = helper.init5Numbers()
    for _i in range(0, len(indexes)):
        rx = indexes[_i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        # 第五位
        w6 = str(ws.cell(row=rx, column=6).value)
        if w2 == "None":
            break
        #
        if isRange == 1:
            _qianRange = list(map(int, configuration['qianRange']))
            _baiRange = list(map(int, configuration['baiRange']))
            _shiRange = list(map(int, configuration['shiRange']))
            _geRange = list(map(int, configuration['geRange']))
            for i in range(0, len(w2)):
                if _qianRange.count(int(w2[i])) <= 0:
                    continue
                for j in range(0, len(w3)):
                    if _baiRange.count(int(w3[j])) <= 0:
                        continue
                    for m in range(0, len(w4)):
                        if _shiRange.count(int(w4[m])) <= 0:
                            continue
                        for n in range(0, len(w5)):
                            if _geRange.count(int(w5[n])) <= 0:
                                continue
                            for k in range(0, len(w6)):
                                numbers[w2[i] + w3[j] + w4[m] +
                                        w5[n]+w6[k]].append(w1)
            pass
        else:
            for i in range(0, len(w2)):

                for j in range(0, len(w3)):

                    for m in range(0, len(w4)):

                        for n in range(0, len(w5)):

                            for k in range(0, len(w6)):
                                index = str(w2[i])+str(w3[j]) + \
                                    str(w4[m])+str(w5[n])+str(w6[k])
                                numbers[index].append(w1)

            pass
    # 00000-99999
    for i in range(0, 100000):
        n = "%05d" % i
        numbers[n].sort()
    #
    numbers = numbers.items()
    result = sorted(numbers, key=lambda i: len(i[1]), reverse=True)
    #
    targetSet = set()
    for item in result:
        memberCount = len(item[1])
        if memberCount >= set_range[0] and memberCount <= set_range[1]:
            targetSet.add(item[0])
            pass
    return targetSet


def readWbFrom4Index(ws, indexes, set_range):
    #
    global configuration
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    isRange = configuration['isRange']
    #
    numbers = helper.init4Numbers()
    for _i in range(0, len(indexes)):
        rx = indexes[_i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        if w2 == "None":
            break
        #
        if isRange == 1:
            _qianRange = list(map(int, configuration['qianRange']))
            _baiRange = list(map(int, configuration['baiRange']))
            _shiRange = list(map(int, configuration['shiRange']))
            _geRange = list(map(int, configuration['geRange']))
            for i in range(0, len(w2)):
                if _qianRange.count(int(w2[i])) <= 0:
                    continue
                for j in range(0, len(w3)):
                    if _baiRange.count(int(w3[j])) <= 0:
                        continue
                    for m in range(0, len(w4)):
                        if _shiRange.count(int(w4[m])) <= 0:
                            continue
                        for n in range(0, len(w5)):
                            if _geRange.count(int(w5[n])) <= 0:
                                continue
                            numbers[w2[i] + w3[j] + w4[m] +
                                    w5[n]].append(w1)
            pass
        else:
            for i in range(0, len(w2)):

                for j in range(0, len(w3)):

                    for m in range(0, len(w4)):

                        for n in range(0, len(w5)):

                            index = str(w2[i])+str(w3[j]) + \
                                str(w4[m])+str(w5[n])
                            numbers[index].append(w1)

            pass
    #
    for i in range(0, 10000):
        n = "%04d" % i
        numbers[n].sort()
    #
    numbers = numbers.items()
    result = sorted(numbers, key=lambda i: len(i[1]), reverse=True)
    #
    targetSet = set()
    for item in result:
        memberCount = len(item[1])
        if memberCount >= set_range[0] and memberCount <= set_range[1]:
            targetSet.add(item[0])
            pass
    return targetSet


def readWbFromIndexSource(ws, indexes, filePath, parentIndex):
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5, 'E': 6}
    parts = [
        dict(onePart=('A', 'B'), twoPart=('C', 'D', 'E')),
        dict(onePart=('A', 'C'), twoPart=('B', 'D', 'E')),
        dict(onePart=('A', 'D'), twoPart=('B', 'C', 'E')),
    ]
    #
    key = 1
    for part in parts:
        result = []
        onePart = part['onePart']
        twoPart = part['twoPart']

        for i in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': '', 'E': ''}
            rx = indexes[i] + 1
            for item in onePart:
                column = columnList[item]
                dataList[item] = (
                    str(ws.cell(row=rx, column=column).value))

            _dataList = dataList.copy()

            for j in range(0, len(indexes)):
                _rx = indexes[j]+1
                for item in twoPart:
                    _column = columnList[item]
                    _dataList[item] = (
                        str(ws.cell(row=_rx, column=_column).value))
                result.append(_dataList)
                _dataList = dataList.copy()

        fullPath = exportSet(filePath, result, 'jiaocha',
                             str(parentIndex)+str(key))
        key += 1
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)


def readWbFrom4IndexSource(ws, indexes, filePath, parentIndex):
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
    parts = [
        dict(onePart=('A', 'B'), twoPart=('C', 'D')),
        dict(onePart=('A', 'C'), twoPart=('B', 'D')),
        dict(onePart=('A', 'D'), twoPart=('B', 'C')),
    ]
    #
    key = 1
    for part in parts:
        result = []
        onePart = part['onePart']
        twoPart = part['twoPart']

        for i in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            rx = indexes[i] + 1
            for item in onePart:
                column = columnList[item]
                dataList[item] = (
                    str(ws.cell(row=rx, column=column).value))

            _dataList = dataList.copy()

            for j in range(0, len(indexes)):
                _rx = indexes[j]+1
                for item in twoPart:
                    _column = columnList[item]
                    _dataList[item] = (
                        str(ws.cell(row=_rx, column=_column).value))
                result.append(_dataList)
                _dataList = dataList.copy()

        fullPath = export4Set(filePath, result, 'jiaocha',
                              str(parentIndex)+str(key))
        key += 1
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)


def exportSet(path, data, bonus='', key=0):
    _workbook = xlwt.Workbook(encoding='utf-8')
    # 样式
    style = xlwt.XFStyle()  # 创建一个样式对象，初始化样式
    al = xlwt.Alignment()
    al.horz = 0x02      # 设置水平居中
    al.vert = 0x01      # 设置垂直居中
    style.alignment = al
    # 注意: 在add_sheet时, 置参数cell_overwrite_ok=True, 可以覆盖原单元格中数据。
    # cell_overwrite_ok默认为False, 覆盖的话, 会抛出异常.
    sheet = _workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
    # 获取并写入数据段信息
    sheet.write(0, 0, '姓名', style)
    sheet.write(0, 1, '数据1', style)
    sheet.write(0, 2, '数据2', style)
    sheet.write(0, 3, '数据3', style)
    sheet.write(0, 4, '数据4', style)
    sheet.write(0, 5, '数据5', style)
    for row in range(0, len(data)):
        for index in range(0, 5):
            sheet.write(row+1, 0, 'A'+str(row+1), style)
            sheet.write(row+1, index+1, int(data[row]
                                            [['A', 'B', 'C', 'D', 'E'][index]]), style)
    #
    if bonus == '':
        fullPath = path+'\\源文件'+'总计'+str(len(data))+'.xls'
    if bonus == 'add':
        fullPath = path+'\\叠加文件'+str(key)+'.xls'
    if bonus == 'leijia':
        fullPath = path+'\\指定数量累加文件'+str(key)+'.xls'
    if bonus == 'leijiadir':
        fullPath = path+'\\文件夹同序累加文件'+str(key)+'.xls'
    if bonus == 'jiaocha':
        fullPath = path+'\\交叉文件'+str(key)+'.xls'
    if bonus == 'zuhe':
        fullPath = path+'\\组合文件'+str(key)+'.xls'
    _workbook.save(fullPath)
    return fullPath


def export4Set(path, data, bonus='', key=0):
    _workbook = xlwt.Workbook(encoding='utf-8')
    # 样式
    style = xlwt.XFStyle()  # 创建一个样式对象，初始化样式
    al = xlwt.Alignment()
    al.horz = 0x02      # 设置水平居中
    al.vert = 0x01      # 设置垂直居中
    style.alignment = al
    # 注意: 在add_sheet时, 置参数cell_overwrite_ok=True, 可以覆盖原单元格中数据。
    # cell_overwrite_ok默认为False, 覆盖的话, 会抛出异常.
    sheet = _workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
    # 获取并写入数据段信息
    sheet.write(0, 0, '姓名', style)
    sheet.write(0, 1, '数据1', style)
    sheet.write(0, 2, '数据2', style)
    sheet.write(0, 3, '数据3', style)
    sheet.write(0, 4, '数据4', style)
    for row in range(0, len(data)):
        for index in range(0, 4):
            sheet.write(row+1, 0, 'A'+str(row+1), style)
            sheet.write(row+1, index+1, int(data[row]
                                            [['A', 'B', 'C', 'D'][index]]), style)
    #
    if bonus == '':
        fullPath = path+'\\源文件'+'总计'+str(len(data))+'.xls'
    if bonus == 'add':
        fullPath = path+'\\叠加文件'+str(key)+'.xls'
    if bonus == 'leijia':
        fullPath = path+'\\指定数量累加文件'+str(key)+'.xls'
    if bonus == 'leijiadir':
        fullPath = path+'\\文件夹同序累加文件'+str(key)+'.xls'
    if bonus == 'jiaocha':
        fullPath = path+'\\交叉文件'+str(key)+'.xls'
    if bonus == 'zuhe':
        fullPath = path+'\\组合文件'+str(key)+'.xls'
    _workbook.save(fullPath)
    return fullPath


# 单纯累加
def readWbFromIndexAdd(wsList, eachFileDataCount, filePath, parentIndex):
    indexes = list(range(1, eachFileDataCount + 1))
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5, 'E': 6}
    parts = ('A', 'B', 'C', 'D', 'E')
    #
    result = []
    for ws in wsList:
        #
        for j in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': '', 'E': ''}
            for item in parts:
                _column = columnList[item]
                _rx = indexes[j]+1
                dataList[item] = (
                    str(ws.cell(row=_rx, column=_column).value))
            result.append(dataList)
        #
    fullPath = exportSet(filePath, result, 'leijia',
                         str(parentIndex))
    #
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fullPath)

    # FileFormat = 51 is for .xlsx extension
    wb.SaveAs(fullPath+"x", FileFormat=51)
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    #
    os.remove(fullPath)


def readWbFromIndex4Add(wsList, eachFileDataCount, filePath, parentIndex):
    indexes = list(range(1, eachFileDataCount + 1))
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
    parts = ('A', 'B', 'C', 'D')
    #
    result = []
    for ws in wsList:
        #
        for j in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            for item in parts:
                _column = columnList[item]
                _rx = indexes[j]+1
                dataList[item] = (
                    str(ws.cell(row=_rx, column=_column).value))
            result.append(dataList)
        #
    fullPath = export4Set(filePath, result, 'leijia',
                          str(parentIndex))
    #
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fullPath)

    # FileFormat = 51 is for .xlsx extension
    wb.SaveAs(fullPath+"x", FileFormat=51)
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    #
    os.remove(fullPath)


# 累加变种
def readWbFromIndexAddDir(wsList, eachFileDataCount, filePath, parentIndex, type):
    if type == 5:
        indexes = list(range(1, eachFileDataCount + 1))
        columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5, 'E': 6}
        parts = ('A', 'B', 'C', 'D', 'E')
    else:
        indexes = list(range(1, eachFileDataCount + 1))
        columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
        parts = ('A', 'B', 'C', 'D')
    #
    result = []
    for ws in wsList:
        #
        for j in range(0, len(indexes)):
            if type == 5:
                dataList = {'A': '', 'B': '', 'C': '', 'D': '', 'E': ''}
            else:
                dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            for item in parts:
                _column = columnList[item]
                _rx = indexes[j]+1
                dataList[item] = (
                    str(ws.cell(row=_rx, column=_column).value))
            result.append(dataList)
        #
    fullPath = export4Set(filePath, result, 'leijia',
                          str(parentIndex))
    #
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fullPath)

    # FileFormat = 51 is for .xlsx extension
    wb.SaveAs(fullPath+"x", FileFormat=51)
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    #
    os.remove(fullPath)


# 交叉累加
def readWbFromIndexAddSource(ws, _ws, sourceCount, filePath, parentIndex):
    indexes = list(range(1, sourceCount*2 + 1))
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5, 'E': 6}
    parts = [
        dict(onePart=('A', 'B'), twoPart=('C', 'D', 'E')),
        dict(onePart=('A', 'C'), twoPart=('B', 'D', 'E')),
        dict(onePart=('A', 'D'), twoPart=('B', 'C', 'E')),
    ]
    #
    key = 1
    for part in parts:
        result = []
        onePart = part['onePart']
        twoPart = part['twoPart']
        #
        for i in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': '', 'E': ''}
            for item in onePart:
                column = columnList[item]
                if i >= sourceCount:
                    rx = indexes[i-sourceCount] + 1
                    dataList[item] = (
                        str(_ws.cell(row=rx, column=column).value))
                else:
                    rx = indexes[i] + 1
                    dataList[item] = (
                        str(ws.cell(row=rx, column=column).value))

            _dataList = dataList.copy()

            for j in range(0, len(indexes)):
                for item in twoPart:
                    _column = columnList[item]
                    if j >= sourceCount:
                        _rx = indexes[j-sourceCount]+1
                        _dataList[item] = (
                            str(_ws.cell(row=_rx, column=_column).value))
                    else:
                        _rx = indexes[j]+1
                        _dataList[item] = (
                            str(ws.cell(row=_rx, column=_column).value))
                result.append(_dataList)
                _dataList = dataList.copy()

        #
        fullPath = exportSet(filePath, result, 'add',
                             str(parentIndex)+str(key))
        key += 1
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)


def readWbFrom4IndexAddSource(ws, _ws, sourceCount, filePath, parentIndex):
    indexes = list(range(1, sourceCount*2 + 1))
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
    parts = [
        dict(onePart=('A', 'B'), twoPart=('C', 'D')),
        dict(onePart=('A', 'C'), twoPart=('B', 'D')),
        dict(onePart=('A', 'D'), twoPart=('B', 'C')),
    ]
    #
    key = 1
    for part in parts:
        result = []
        onePart = part['onePart']
        twoPart = part['twoPart']
        #
        for i in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            for item in onePart:
                column = columnList[item]
                if i >= sourceCount:
                    rx = indexes[i-sourceCount] + 1
                    dataList[item] = (
                        str(_ws.cell(row=rx, column=column).value))
                else:
                    rx = indexes[i] + 1
                    dataList[item] = (
                        str(ws.cell(row=rx, column=column).value))

            _dataList = dataList.copy()

            for j in range(0, len(indexes)):
                for item in twoPart:
                    _column = columnList[item]
                    if j >= sourceCount:
                        _rx = indexes[j-sourceCount]+1
                        _dataList[item] = (
                            str(_ws.cell(row=_rx, column=_column).value))
                    else:
                        _rx = indexes[j]+1
                        _dataList[item] = (
                            str(ws.cell(row=_rx, column=_column).value))
                result.append(_dataList)
                _dataList = dataList.copy()

        #
        fullPath = exportSet(filePath, result, 'add',
                             str(parentIndex)+str(key))
        key += 1
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)


# 命中情况统计 (45通用)
def missingEachWbFromIndex(ws, indexes, rightNumber):
    #
    count = 0
    for _i in range(0, len(indexes)):
        rx = indexes[_i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        # 第五位
        w6 = str(ws.cell(row=rx, column=6).value)
        if w2 == "None":
            break
        #
        for i in range(0, len(w2)):
            if w2[i] == rightNumber[0]:
                for j in range(0, len(w3)):
                    if w3[j] == rightNumber[1]:
                        for m in range(0, len(w4)):
                            if w4[m] == rightNumber[2]:
                                for n in range(0, len(w5)):
                                    if w5[n] == rightNumber[3]:
                                        if len(rightNumber) == 4:
                                            count += 1
                                            pass
                                        else:
                                            for k in range(0, len(w6)):
                                                if w6[k] == rightNumber[4]:
                                                    count += 1

            pass
    return count


# 批量excel单个随机组合 每个生成对应文件夹 （仅4）
def mutilOriginPaste(path, filePath, totalCount, fileCount, dataCount, index):
    wb = openpyxl.load_workbook(path)
    sheetsNames = wb.sheetnames
    ws = wb[sheetsNames[0]]
    indexes = range(1, totalCount+1)
    aList = []
    bList = []
    cList = []
    dList = []
    # 获取excel abcd 4列总数据
    for i in range(0, len(indexes)):
        rx = indexes[i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        if w2 == "None":
            break
        else:
            aList.append(w2)
            bList.append(w3)
            cList.append(w4)
            dList.append(w5)
    #
    pathList = []
    for i in range(0, fileCount):
        #
        result = []
        for k in range(0, dataCount):
            indexList = random.sample(range(0, totalCount), 5)
            result.append(dict(A=aList[indexList[0]], B=bList[indexList[1]],
                               C=cList[indexList[2]], D=dList[indexList[3]]))
        #
        if index == 0:
            fullPath = export4Set(filePath, result, 'zuhe',
                                  str(i+1))
            pass
        else:
            directoryExist = os.path.exists(filePath+'\\组合文件'+str(index))
            if directoryExist == False:
                os.makedirs(filePath+'\\组合文件'+str(index))
            #
            fullPath = export4Set(filePath+'\\组合文件'+str(index), result, 'zuhe',
                                  str(i+1))
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)
        #
        pathList.append(fullPath+'x')
    return pathList


# 多线程批量进行单区间处理 （仅4）
def partsReadWbFromIndex(pathArray, eachDataCount, section, filePath):
    # 线程列表
    pList = []
    # 需要线程数
    times = len(pathArray)
    for i in range(0, times):
        pathList = pathArray[i]
        p = mp.Process(target=readWbFromIndexThread, args=(
            pathList, eachDataCount, section, filePath, str(i+1)))
        p.start()
        pList.append(p)
        pass
    #
    for i in range(0, len(pList)):
        pList[i].join()
    pass


def readWbFromIndexThread(pathList, eachDataCount, section, filePath, key):
    #
    directoryExist = os.path.exists(filePath+'\\组合文件夹'+str(key))
    if directoryExist == False:
        os.makedirs(filePath+'\\组合文件夹'+str(key))
    #
    index = 1
    targetSet = {}
    for path in pathList:
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, eachDataCount+1)
        resultSet = readWbFrom4Index(
            ws, indexes, section)
        # targetSet
        if len(targetSet) == 0:
            targetSet = resultSet
        else:
            targetSet = targetSet & resultSet
        #
        file = open(filePath+'\\组合文件夹'+str(key)+"\\批量结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        index += 1
    #
    setFile = open(filePath+'\\组合文件夹'+str(key)+"\\结果交集_"+str(key)+"_" +
                   str(len(targetSet))+'.txt', 'w')
    setFile.write(','.join(list(targetSet)))
    pass
