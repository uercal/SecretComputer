from fileinput import filename
from math import trunc
import sys
from tkinter.constants import FALSE, NONE
from typing import ItemsView
from PyQt5.QtCore import right
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QTableWidgetItem
from tkinter import Tk, filedialog
from easygui import fileopenbox, diropenbox, ccbox, enterbox, passwordbox,multenterbox
import os
import configparser
import json
import random
import time
import openpyxl
#
from bs4 import BeautifulSoup
import urllib.request
from openpyxl.descriptors.base import Alias
import requests
#
import win32com.client as win32
import xlwt
#
from itertools import combinations, count, permutations
#
import window
#
import config
#
import helper
import excel5Helper as excelHelper
import txtHelper

# 基础类


def basicTxtSet(method):
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('选择第一个数据')
    target = filedialog.askopenfilename(
        filetypes=[('text files', '.txt')], title="选择第一个数据")
    if target == '':
        ui.textBrowser.append('取消')
        return
    ui.textBrowser.append('选择第二个数据')
    before = filedialog.askopenfilename(
        filetypes=[('text files', '.txt')], title="选择第二个数据")
    if before == '':
        ui.textBrowser.append('取消')
        return
    targetSet = set(open(target, "r").read().split(','))
    beforeSet = set(open(before, "r").read().split(','))
    resultSet = {
        'jiao': targetSet & beforeSet,
        'cha': targetSet - beforeSet,
        'bing': targetSet | beforeSet,
    }[method]
    resultStr = {
        'jiao': 'txt交集',
        'cha': 'txt差集',
        'bing': 'txt并集',
    }[method]
    ui.textBrowser.append('选择结果存放目录')
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)

    # 位数筛选 开启判定
    if configuration['isRange'] == 1:
        resultSet = txtHelper.positionRangeFilter(resultSet)

    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & resultSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''

    file = open(filePath+'\\'+resultStr+'_' +
                str(len(resultSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(resultSet)))
    file.flush()
    file.close()
    ui.textBrowser.append('运算结果，'+resultStr+'数量：'+str(len(resultSet)))
    ui.textBrowser.append(
        '================================================')
    os.startfile(filePath)
    # postData(filePath+'\\'+resultStr+'_'+str(len(resultSet))+'.txt')


def jiaoji():
    basicTxtSet('jiao')


def chaji():
    basicTxtSet('cha')


def bingji():
    basicTxtSet('bing')


def check():
    root = Tk()
    root.withdraw()
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    resultSet = set()
    for path in cur:
        f = open(path, "r")  # 设置文件对象
        tSet = set(f.read().split(','))
        if len(resultSet) == 0:
            resultSet = tSet
            pass
        else:
            resultSet = resultSet.intersection(tSet)
        pass
    fname = filedialog.asksaveasfilename(
        title='保存文件', filetypes=[('text files', '.txt')])
    if fname == '':
        ui.textBrowser.append('取消')
        return

     # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)

    # 位数筛选 开启判定
    if configuration['isRange'] == 1:
        resultSet = txtHelper.positionRangeFilter(resultSet)

    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & resultSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''
    file = open(fname+'总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(resultSet)))
    file.flush()
    file.close()
    root.destroy()
    root.mainloop()
    os.startfile(fname+'总计'+str(len(resultSet))+bonusStr+'.txt')
    ui.textBrowser.append('运算结果 : 数量：'+str(len(resultSet)))
    ui.textBrowser.append(
        '================================================')
    # postData(fname+'总计'+str(len(resultSet))+'.txt')


# 多txt求并 叠加
def mutilBind():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('叠加集合txt。。。。')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    resultSet = set()
    for path in cur:
        f = open(path, "r")  # 设置文件对象
        tSet = set(f.read().split(','))
        if len(resultSet) == 0:
            resultSet = tSet
            pass
        else:
            resultSet = resultSet.union(tSet)
        pass
    fname = filedialog.asksaveasfilename(
        title='保存文件', filetypes=[('text files', '.txt')])
    if fname == '':
        ui.textBrowser.append('取消')
        return
     # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    # 位数筛选 开启判定
    if configuration['isRange'] == 1:
        resultSet = txtHelper.positionRangeFilter(resultSet)

    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & resultSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''

    file = open(fname+'总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(resultSet)))
    file.flush()
    file.close()
    root.destroy()
    root.mainloop()
    os.startfile(fname+'总计'+str(len(resultSet))+bonusStr+'.txt')
    # postData(fname+'总计'+str(len(resultSet))+'.txt')


# 多txt求差 目标
def mutilCha():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('批量目标差集合txt。。。。')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
     # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    completedSet = helper.init5NumbersSet()
    index = 1
    for path in cur:
        f = open(path, "r")  # 设置文件对象
        tSet = set(f.read().split(','))
        resultSet = completedSet - tSet
        #
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\目标差集结果集'+str(index) +
                    '_'+str(len(resultSet))+bonusStr+'.txt', 'w')
        index += 1
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        pass
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('各txt差集计算完成！')
    os.startfile(filePath)


# txt 逆算
def reverseCheck():
    root = Tk()
    root.withdraw()
    aPosition = {
        '0': 0, '1': 0, '2': 0, '3': 0, '4': 0, '5': 0, '6': 0, '7': 0, '8': 0, '9': 0
    }
    bPosition = {
        '0': 0, '1': 0, '2': 0, '3': 0, '4': 0, '5': 0, '6': 0, '7': 0, '8': 0, '9': 0
    }
    cPosition = {
        '0': 0, '1': 0, '2': 0, '3': 0, '4': 0, '5': 0, '6': 0, '7': 0, '8': 0, '9': 0
    }
    dPosition = {
        '0': 0, '1': 0, '2': 0, '3': 0, '4': 0, '5': 0, '6': 0, '7': 0, '8': 0, '9': 0
    }
    ePosition = {
        '0': 0, '1': 0, '2': 0, '3': 0, '4': 0, '5': 0, '6': 0, '7': 0, '8': 0, '9': 0
    }
    position = [
        aPosition, bPosition, cPosition, dPosition, ePosition
    ]
    target = filedialog.askopenfilename(
        filetypes=[('text files', '.txt')], title="选择第一个数据")
    if target == '':
        ui.textBrowser.append('取消')
        return False
    targetList = open(target, "r").read().split(',')
    for i in range(0, len(targetList)):
        itemStr = targetList[i]
        for j in range(0, len(itemStr)):
            position[j][str(itemStr[j])] += 1

    for i in range(0, len(position)):
        res = sorted(position[i].items(),
                     key=lambda item: item[1], reverse=True)
        ui.textBrowser.append('第'+str(i+1)+'位置结果：')
        recommon = ''
        for j in range(0, len(res)):
            uiLabel = '数字 '+res[j][0]+' 次数:'
            ui.textBrowser.append(uiLabel+str(res[j][1]))
            #
            recommon += res[j][0]
        ui.textBrowser.append('6位数推荐结果：'+recommon[0:6])
        ui.textBrowser.append('7位数推荐结果：'+recommon[0:7])
        ui.textBrowser.append('========== 分割线 ============')
    pass


# txt 批量随机抽取txt
def randomMutilTxt():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('批量抽取txt.........')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    # 输入数量
    fileCount = int(enterbox("随机取多少数据进行交集?", '确认', "1"))
    if fileCount <= 0:
        ui.textBrowser.append('交集数据数不能为0')
        return
    handleCount = int(enterbox("目标交集数量?", '确认', "0"))
    # 判断命中
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber'] if isCheckRight == True else 0
    # completedSet = helper.init4NumbersSet if version == True else helper.init5NumbersSet()
    # 结果目录
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    cur = list(cur)
    # 获得抽取后的path数组
    allPickList = list(combinations(cur, fileCount))
    #
    if handleCount == 0:
        ui.textBrowser.append('即将输出所有结果('+str(len(allPickList))+')')
        handleCount = len(allPickList)

    for i in range(0, handleCount):
        pathList = random.choice(allPickList)
        allPickList.remove(pathList)
        # handler
        labelList = []
        for m in range(0, len(pathList)):
            lastP = pathList[m].rfind('/')+1
            labelList.append(pathList[m][lastP:-4])
        ui.textBrowser.append(
            '抽取结果'+str(i+1)+':['+' , '.join(labelList)+']')
        txtHelper.mutilTxtCheck(
            pathList, filePath, '随机抽取交集'+str(i+1), rightNumber, isRange=configuration['isRange'])
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成！')
    os.startfile(filePath)
    pass

#  multiPositionHandler
def multiPositionHandler():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('选取txt文件.........')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return            
    if len(cur) > 2:
        ui.textBrowser.append('最多只能选取1-2个文件')
        return 
    # 判断命中
    type = ccbox('处理算法', '提示', ('交集', '并集'))
    handleType = 'intersect' if type == True else 'union'
    # 结果目录
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return    
    if len(cur) == 1 :
        handleNumber = int(enterbox("多少个数据进行处理（不能低于2）?", '确认', "2"))
    else:
        handleNumber = 0   
    c = txtHelper.filesSetHandler(cur,filePath,handleType,handleNumber)
    if c == False:
        return c
    ui.textBrowser.append('计算完成！')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    pass


# findEmptyHandler
def findEmptyHandler():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('选取txt文件.........')
    cur = filedialog.askopenfilename(filetypes=[('text files', '.txt')])
    # 结果目录
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return        
    handleNumber = enterbox("多少个数据进行处理（不能低于3）?", '确认', "3")
    if handleNumber == None:
        return 
    handleNumber = int(handleNumber)
    fileName = txtHelper.findEmptySet(cur,filePath,handleNumber)
    ui.textBrowser.append(fileName)
    ui.textBrowser.append('计算完成！')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    pass

# singleNumberRate
def singleNumberRate():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('选取txt文件.........')
    cur = filedialog.askopenfilename(filetypes=[('text files', '.txt')])         
    # 结果目录
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return     
    counterList = txtHelper.singleNumberRate(cur,filePath)
    ui.textBrowser.append('计算完成！')
    for i in range(0,len(counterList)):
        ui.textBrowser.append(counterList[i])                    
    os.startfile(filePath)
    root.destroy()
    root.mainloop()    
    pass


# 单双重 数位限制
def danshuangPosition():
    positionType = ccbox('选择类型', '提示', ('单重', '双重'))    
    if positionType == True:
        file = 'position-dan.txt'
    else:
        file = 'position-shuang.txt'
    with open(file,'r') as f:
        positionSet = set(f.read().split(','))
    # 
    actionType = ccbox('输入类型', '提示', ('导入复式txt', '输入位数'))
    if actionType == True:
        root = Tk()
        root.withdraw()        
        cur = filedialog.askopenfilename(filetypes=[('text files', '.txt')])   
        with open(cur,'r') as f:
            targetList = list(f.read().split(','))
    else:
        fieldNames = ["千位", "百位", "十位", "各位"]
        targetList = multenterbox("输入各个位限制数据","位数限制",fieldNames)
        if targetList == None or len(targetList) != 4:
            ui.textBrowser.append('位数输入错误')
            return False            
        else:
            targetList = [','.join(targetList)]
    # action txt export    
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    txtHelper.positionNumberInterset(targetList,filePath,positionSet,actionType,positionType)
    ui.textBrowser.append('计算完成')    
    pass

# 数位限制
def actionpositionCountOrder():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('选取txt文件.........')
    cur = filedialog.askopenfilename(filetypes=[('text file', '.txt')])      
    if cur == None:
        return 
    # 结果目录
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return     
    txtHelper.positionTxtCountHandler(cur,filePath)
    ui.textBrowser.append('计算完成！')                  
    os.startfile(filePath)
    root.destroy()
    root.mainloop()    
    pass
    pass

# 序号文件 分4组 求交并 限制
def actionorderFileCombine():
    root = Tk()
    root.withdraw()
    actionType = ccbox('输入类型', '提示', ('原始数据', '交空数据'))
    ui.textBrowser.append('选取txt文件.........')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])      
    if cur == None:
        return 
    # 
    handleNumber = int(enterbox("需要多少个交集（4的倍数，每4个合并为1个并集）?", '确认', "4"))
    if handleNumber == None or handleNumber%4 != 0:
        ui.textBrowser.append('参数错误！')  
        return 
    # 
    # 结果目录
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return     
    txtHelper.orderFileInterBind(cur,filePath,handleNumber,actionType)
    ui.textBrowser.append('计算完成！')                  
    os.startfile(filePath)
    root.destroy()
    root.mainloop()    
    pass


# 求空原始文件 前五 后五 交并
def fiveEmptyBindInter():
    import os
    root = Tk()
    root.withdraw()    
    # 
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return     
    txtHelper.testDemo(filePath)    
    os.startfile(filePath)
    root.destroy()
    root.mainloop()    

    return


    #     
    times = 0
    total = []
    root = Tk()
    root.withdraw()    
    ui.textBrowser.append('选取文件夹.........')
    cate = filedialog.askdirectory()    
    _totalCount = 1
    
    for root,dirs,files in os.walk(cate): 
        for dir in dirs: 
            category = os.path.join(root,dir)            
            for _root,_dirs,_files in os.walk(category):
                cur = [os.path.join(_root,c) for c in _files]  
                # ACTION
                right = (cur[0].split('\\')[1].split(' ')[-1])                                    
                # rightNumber = enterbox("依次输入 头千百十 ", '确认', right)
                # handleCount = enterbox("输入运行次数 ", '确认', "30")
                rightNumber = right
                handleCount = 20
                if rightNumber == None or handleCount == None:
                    return
                # 
                allBoolean = True                
                for curPath in cur:
                    # [namePath,checkNumber,boolean,resultBind] = txtHelper.fiveEmptyBindInterHandlerSingle(curPath,rightNumber,handleCount)        
                    # ui.textBrowser.append(namePath+'('+checkNumber+')['+','.join(resultBind)+']:'+'命中'if boolean else namePath+'('+checkNumber+')['+','.join(resultBind)+']:'+'落空')
                    # if boolean == False:
                    #     allBoolean = False    
                    [namePath,resultInter,boolean,count] = txtHelper.fiveEmptyBindInterHandlerSingle(curPath,rightNumber,handleCount)
                    _totalCount = _totalCount * count
                    ui.textBrowser.append(namePath+'['+','.join(resultInter)+']:不符合(×包含中奖位置)'if boolean else namePath+'['+','.join(resultInter)+']:符合（√不包含中奖位置）')
                    if boolean == True:
                        allBoolean = False
                # ui.textBrowser.append('计算结果：'+'全部命中' if allBoolean else '落空')            
                if allBoolean == True:
                    times = times+1
                    total.append(_totalCount)
                    ui.textBrowser.append('_totalCount:'+str(_totalCount))                                          
                ui.textBrowser.append('计算结果：'+'全部命中' if allBoolean else '落空')            
                ui.textBrowser.append('times:'+str(times))
                _totalCount = 1  
                if len(total) > 0:
                    avgCount = sum(total)/len(total)
                    ui.textBrowser.append('avgCount:'+str(avgCount))                     
                #     
    # if cur == None or cur =='':
    #     return                 
    pass


# 
def shaCodeFullCha():
    from collections import Counter
    root = Tk()
    root.withdraw()
    cur = filedialog.askopenfilename(filetypes=[('text files', '.txt')])
    # 结果目录
    filePath = diropenbox('结果存放目录')
    # action handler
    f = open(cur,'r')
    fList = f.read().split('\n')
    valueList = []
    totalList = {}
    indexSet = {}
    for item in fList:
        [index,value] = item.split(':')
        totalList[value] = index
        valueList.append(value)
        indexSet[index] = 0
    # 
    group = list(combinations(valueList, 4))
    result = {'0':0,'1':0,'2':0,'3':0,'4':0,'5':0,'6':0,'7':0,'8':0,'9':0}
    
    resultList = []
    indexStrList = []
    for item in group:
        item = [a for a in item]
        bindItem = helper.listForBind(item)
        bindItem = [int(b) for b in bindItem]
        resultItem = list(set(range(0,10)) - set(bindItem))
        if len(resultItem)>0:
            listResult = [str(c) for c in resultItem]
            for j in listResult:
                result[j] = result[j]+1
            resultStr = ','.join(listResult)
            resultList.append(resultStr)
        if len(resultItem)==0:
            indexStr = ''
            for j in item:
                indexStr = indexStr+totalList[j]
                indexSet[totalList[j]] = indexSet[totalList[j]]+1
            indexStrList.append(indexStr)
    # 
    counterRate = Counter(result).most_common()
    counterList = []
    for k in range(0,len(counterRate)):
        numStr = str(counterRate[k][0])
        countStr = str(counterRate[k][1])        
        counterList.append(numStr+'（'+countStr+'）')    
    # 
    counterIndexRate = Counter(indexSet).most_common()
    counterIndexList = []
    for k in range(0,len(counterIndexRate)):
        _numStr = str(counterIndexRate[k][0])
        _countStr = str(counterIndexRate[k][1])        
        counterIndexList.append(_numStr+'（'+_countStr+'）')   

    __file = open(filePath+'\\'+'满集结果组合.txt', 'w')
    __file.write('\n'.join(indexStrList))
    __file.flush()
    __file.close()
    ___file = open(filePath+'\\'+'满集结果频次.txt', 'w')
    ___file.write('\n'.join(counterIndexList))
    ___file.flush()
    ___file.close()
    # 
    _file = open(filePath+'\\'+'测试差集结果频次.txt', 'w')
    _file.write('\n'.join(counterList))
    _file.flush()
    _file.close()
    # 
    file = open(filePath+'\\'+'测试差集结果.txt', 'w')
    file.write('\n'.join(resultList))
    file.flush()
    file.close()
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    pass



# 拓展类（爬虫
def climpPage(page):
    ui.textBrowser.append("爬取第"+str(page)+"页数据中。。。。")
    url = 'http://caipiao.eastmoney.com/pub/Result/History/pl5?page='+str(page)
    html = requests.get(url)
    html.encoding = 'UTF-8'
    html = html.text
    soup = BeautifulSoup(html, "lxml")
    tag = soup.findAll("span", class_="pellet pellet-primary pellet-sm red")
    numbers = []
    string = []
    for i in range(0, len(tag)):
        if len(string) == 5:
            numbers.append(''.join(string))
            string = []
            string.append(tag[i].string)
        else:
            string.append(tag[i].string)
    return numbers


def exportList2Txt(filename, dataList):
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    file = open(filePath+'\\'+filename+'_' +
                str(len(dataList))+'.txt', 'w')
    file.write(','.join(list(dataList)))
    file.flush()
    file.close()
    os.startfile(filePath)
# 输出文本


def exportTxt(filename, dataSet):
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & dataSet) > 0:
            bonusStr = '_命中'
    file = open(filePath+'\\'+filename+'_' +
                str(len(dataSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(dataSet)))
    file.flush()
    file.close()
    # postData(filePath+'\\'+filename+'_'+str(len(dataSet))+'.txt')
    os.startfile(filePath)
# 输出文本


def exportTxt3(filePath, filename, dataStr):
    #
    dataSet = set(dataStr.split(','))
    if len({rightNumber} & dataSet) > 0:
        is_exist = '落入'
    else:
        is_exist = ''
    file = open(filePath+'\\'+filename+'_'+is_exist +
                '_'+str(len(dataSet))+'.txt', 'w')
    file.write(dataStr)
    file.flush()
    file.close()
    # postData(filePath+'\\'+filename+'_'+is_exist +
    #          '_'+str(len(dataSet))+'.txt')

# 随机算法


def chouyang(a, n, group, type):
    r = list()
    if type == 'pick':
        while len(r) < group:
            b = random.sample(a, n)
            bStr = ','.join('%s' % id for id in b)
            ui.textBrowser.append('抽取抽样索引值: '+bStr)
            b.sort()  # 排序
            r.append(b)
        return r
    if type == 'group':
        p = True
        while p:
            if len(a) <= n:
                r.append(a)
                return r
            b = random.sample(a, n)
            bStr = ','.join('%s' % id for id in b)
            ui.textBrowser.append('分组抽样索引值: '+bStr)
            b.sort()  # 排序
            r.append(b)
            a = list(set(a).difference(set(b)))  # 去除已抽样的数据
            if len(a) > 0:
                p = True
            else:
                p = False
        return r


def txtRandomInter(indexList, txtRandomCount, txtRandomGroup, cur, type, returnType='intersection'):
    allIndexList = chouyang(indexList, txtRandomCount, txtRandomGroup, type)
    bindSetList = []
    for i in range(0, len(allIndexList)):
        bindSet = set()
        for j in range(0, len(allIndexList[i])):
            index = allIndexList[i][j]
            f = open(cur[index], "r")
            tSet = set(f.read().split(','))
            bindSet = bindSet | tSet
        bindSetList.append(bindSet)
    #
    if returnType == 'bind':
        return bindSetList
    else:
        resultSet = set()
        for i in range(0, len(bindSetList)):
            if i == 0:
                resultSet = bindSetList[i]
            else:
                resultSet = resultSet & bindSetList[i]
        return resultSet


# 随机类
def txtRandomTimes():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机抽取，并计算.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))

    for i in range(0, txtRandomHandlerCount):
        resultSet = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'pick')
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\批量随机排列抽取交集结果'+str(i+1) +
                    '总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtRandomTimesGroup():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机排列分组，并计算.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))

    for i in range(0, txtRandomHandlerCount):
        resultSet = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'group')
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\随机排列分组交集结果'+str(i+1) +
                    '总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtRandomPickBind():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机抽取，并计算所有并集.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    for i in range(0, txtRandomHandlerCount):
        resultSetList = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'pick', 'bind')
        for j in range(0, len(resultSetList)):
            resultSet = resultSetList[j]
            file = open(filePath+'\\随机抽取并集结果'+str(i+1) + '_'+str(j+1) +
                        '总计'+str(len(resultSet))+'.txt', 'w')
            file.write(','.join(list(resultSet)))
            file.flush()
            file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtDiyPickGroupInter():
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始计算.......')
    curDir = diropenbox('目标目录')
    if curDir == None:
        return False
    cur = []
    for _root, dirs, files in os.walk(curDir):
        for file in files:
            if os.path.splitext(file)[1] == '.txt':
                cur.append(os.path.join(_root, file))
    if len(cur) == 0:
        ui.textBrowser.append('该目录下 不存在txt文件')
        return False
    totalCount = len(cur)
    #
    indexList = list(range(0, totalCount))
    #
    handleCount = int(enterbox("重复运算多少次?", '确认', "0"))
    randomCount = int(enterbox("随机多少数据进行一次交集?", '确认', "0"))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    # handler
    for i in range(0, handleCount):
        indexesList = random.sample(indexList, randomCount)
        resultSet = set()
        for j in range(0, len(indexesList)):
            index = indexesList[j]
            f = open(cur[index], "r")
            tSet = set(f.read().split(','))
            if len(resultSet) == 0:
                resultSet = tSet
            else:
                resultSet = resultSet.intersection(tSet)
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\自定义抽取交集结果_'+str(i+1) +
                    '总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtRandomGroupBind():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机分组，并计算所有并集.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    for i in range(0, txtRandomHandlerCount):
        resultSetList = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'group', 'bind')
        for j in range(0, len(resultSetList)):
            resultSet = resultSetList[j]
            file = open(filePath+'\\随机分组并集结果'+str(i+1) + '_'+str(j+1) +
                        '总计'+str(len(resultSet))+'.txt', 'w')
            file.write(','.join(list(resultSet)))
            file.flush()
            file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def postData(filePath):
    try:
        config = configparser.ConfigParser()
        config.read('./url.ini')
        url = config['DEFAULT']['URL']
        data = {"token": "yogo"}
        files = {'file': open(filePath, 'rb')}
        res = requests.post(url=url, files=files, data=data)
    except:
        print('------')
        return False
    else:
        print('---')
        return res


def climb():
    initUrl = 'http://caipiao.eastmoney.com/pub/Result/History/pl5?page=1'
    html = requests.get(initUrl)
    html.encoding = 'UTF-8'
    html = html.text
    soup = BeautifulSoup(html, 'html.parser')
    pageTag = soup.select("div .pagination > ul > a")
    maxPage = 0
    for aItem in pageTag:
        page = 0
        try:
            page = int(aItem.string)
        except ValueError:
            print('...')
        else:
            if page > maxPage:
                maxPage = page
    totalList = []
    ideaCount = 0
    inputMaxPage = int(enterbox("输入最大页数?（最大"+str(maxPage)+"）", '确认', "20"))
    for i in range(1, inputMaxPage+1):
        numbers = climpPage(i)
        for item in numbers:
            ideaCount += 1
            totalList.append(item)
    exportList2Txt('爬取结果', dataList=totalList)


def loadRecent():
    url = 'http://caipiao.eastmoney.com/pub/Result/History/pl5?page=1'
    html = requests.get(url)
    html.encoding = 'UTF-8'
    html = html.text
    soup = BeautifulSoup(html, "lxml")
    timeTd = soup.findAll("td")
    timeList = []
    for i in range(0, len(timeTd)):
        if len(timeList) == 7:
            break
        if i % 8 == 1:
            timeList.append(timeTd[i].string)
    tag = soup.findAll("span", class_="pellet pellet-primary pellet-sm red")
    numbers = []
    string = []
    for i in range(0, len(tag)):
        if len(numbers) == 7:
            break
        if len(string) == 5:
            numbers.append(''.join(string))
            string = []
            string.append(tag[i].string)
        else:
            string.append(tag[i].string)
    pass

    for i in range(0, len(numbers)):
        ui.textBrowser.append(
            "<font size='14'>"+timeList[i]+":<font color='red'>"+numbers[i]+'</font></font>')

# config


def showConfig():
    #
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    configui.rightNumber.setText(configuration['rightNumber'])
    configui.totalCount.setText(configuration['totalCount'])
    configui.hundredCount.setText(configuration['hundredCount'])
    configui.fiftyCount.setText(configuration['fiftyCount'])
    configui.insideCount.setText(configuration['insideCount'])
    configui.handlerCount.setText(configuration['handlerCount'])
    configui.isSingle.setChecked(
        True if configuration['isSingle'] == 1 else False)
    configui.isRange.setChecked(
        True if configuration['isRange'] == 1 else False)
    # qian
    configui.tableWidget.item(0, 0).setText(
        configuration['qianRange'] if 'qianRange' in configuration else '')
    # bai
    configui.tableWidget.item(0, 1).setText(
        configuration['baiRange'] if 'baiRange' in configuration else '')
    # shi
    configui.tableWidget.item(0, 2).setText(
        configuration['shiRange'] if 'shiRange' in configuration else '')
    # ge
    configui.tableWidget.item(0, 3).setText(
        configuration['geRange'] if 'geRange' in configuration else '')
    #
    configui.sourceSection.setText(
        configuration['sourceSection'] if 'sourceSection' in configuration else '')
    configui.sourceCount.setText(
        configuration['sourceCount'] if 'sourceCount' in configuration else '')
    #
    configui.excelSection.setText(
        configuration['excelSection'] if 'excelSection' in configuration else '')
    configui.excelCount.setText(
        configuration['excelCount'] if 'excelCount' in configuration else '')
    configui.leftSection.setText(
        configuration['leftSection'] if 'leftSection' in configuration else '')
    configui.rightSection.setText(
        configuration['rightSection'] if 'rightSection' in configuration else '')
    #
    configui.lineTxtRandom.setText(
        configuration['txtRandomCount'] if 'txtRandomCount' in configuration else '')
    configui.lineTxtRandomGroup.setText(
        configuration['txtRandomGroup'] if 'txtRandomGroup' in configuration else '')
    configui.lineTxtRandomCount.setText(
        configuration['txtRandomHandlerCount'] if 'txtRandomHandlerCount' in configuration else '')
    # sectionArray
    tableWidget_2 = configui.tableWidget_2
    setArray = configuration['setArray']
    tableWidget_2.setRowCount(len(setArray))

    for i in range(0, len(setArray)):
        if tableWidget_2.item(i, 0) == None:
            item1 = QTableWidgetItem(str(setArray[i][0]).strip('[').strip(']'))
            item2 = QTableWidgetItem(str(setArray[i][1]).strip('[').strip(']'))
            item3 = QTableWidgetItem(str(setArray[i][2]).strip('[').strip(']'))
            item1.setTextAlignment(0x0080 | 0x0004)
            item2.setTextAlignment(0x0080 | 0x0004)
            item3.setTextAlignment(0x0080 | 0x0004)
            tableWidget_2.setItem(i, 0, item1)
            tableWidget_2.setItem(i, 1, item2)
            tableWidget_2.setItem(i, 2, item3)
        else:
            tableWidget_2.item(i, 0).setText(
                str(setArray[i][0]).strip('[').strip(']'))
            tableWidget_2.item(i, 1).setText(
                str(setArray[i][1]).strip('[').strip(']'))
            tableWidget_2.item(i, 2).setText(
                str(setArray[i][2]).strip('[').strip(']'))

    dialogWindow.show()


def confirmConfig():
    rightNumber = configui.rightNumber.text()
    totalCount = configui.totalCount.text()
    hundredCount = configui.hundredCount.text()
    fiftyCount = configui.fiftyCount.text()
    insideCount = configui.insideCount.text()
    handlerCount = configui.handlerCount.text()
    isSingle = 1 if configui.isSingle.isChecked() == True else 0
    isRange = 1 if configui.isRange.isChecked() == True else 0
    qianRange = configui.tableWidget.item(0, 0).text()
    baiRange = configui.tableWidget.item(0, 1).text()
    shiRange = configui.tableWidget.item(0, 2).text()
    geRange = configui.tableWidget.item(0, 3).text()
    sourceSection = configui.sourceSection.text()
    sourceCount = configui.sourceCount.text()
    excelSection = configui.excelSection.text()
    excelCount = configui.excelCount.text()
    leftSection = configui.leftSection.text()
    rightSection = configui.rightSection.text()
    txtRandomCount = configui.lineTxtRandom.text()
    txtRandomGroup = configui.lineTxtRandomGroup.text()
    txtRandomHandlerCount = configui.lineTxtRandomCount.text()
    # section
    tableWidget_2 = configui.tableWidget_2
    setArray = []
    for i in range(0, tableWidget_2.rowCount()):
        totalSection = list(
            map(int, tableWidget_2.item(i, 0).text().split(',')))
        hundredSection = list(
            map(int, tableWidget_2.item(i, 1).text().split(',')))
        fiftySection = list(
            map(int, tableWidget_2.item(i, 2).text().split(',')))
        setArray.append([totalSection, hundredSection, fiftySection])

    configuration = dict(rightNumber=rightNumber, totalCount=totalCount, hundredCount=hundredCount, fiftyCount=fiftyCount,
                         insideCount=insideCount, handlerCount=handlerCount, isSingle=isSingle, isRange=isRange,
                         qianRange=qianRange, baiRange=baiRange, shiRange=shiRange, geRange=geRange, setArray=setArray,
                         sourceSection=sourceSection, sourceCount=sourceCount,
                         excelSection=excelSection, excelCount=excelCount, leftSection=leftSection, rightSection=rightSection,
                         txtRandomCount=txtRandomCount, txtRandomGroup=txtRandomGroup, txtRandomHandlerCount=txtRandomHandlerCount)
    jsonObj = json.dumps(configuration)
    with open("setting.json", "a") as f:
        f.seek(0)
        f.truncate()
        f.write(jsonObj)


def addSection():
    self = configui.tableWidget_2
    self.insertRow(self.rowCount())


def delSection():
    self = configui.tableWidget_2
    if self.rowCount() > 1:
        self.removeRow(self.rowCount()-1)


# mainCheck
def mainCheck():
    global configuration
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    global rightNumber
    global totalCount
    global hundredCount
    global fiftyCount
    global handlerCount
    global insideCount
    global isSingle
    global isRange
    global qianRange
    global baiRange
    global shiRange
    global geRange
    global setArray
    rightNumber = configuration['rightNumber']
    totalCount = int(configuration['totalCount'])
    hundredCount = int(configuration['hundredCount'])
    fiftyCount = int(configuration['fiftyCount'])
    handlerCount = int(configuration['handlerCount'])
    insideCount = int(configuration['insideCount'])
    isSingle = configuration['isSingle']
    isRange = configuration['isRange']
    qianRange = configuration['qianRange']
    baiRange = configuration['baiRange']
    shiRange = configuration['shiRange']
    geRange = configuration['geRange']
    setArray = configuration['setArray']
    #
    file = fileopenbox("文件选择", "读取文件")
    if file == None:
        return
    #
    totalResult = []
    for i in range(0, len(setArray)):
        sectionArray = setArray[i]
        insideResult = {"right_set": 0, "final_set": 0, "bind_set": 0}
        allResult = {"right_set": [], "final_set": [], "bind_set": []}
        allLength = {"right_set": 0, "final_set": 0, "bind_set": 0}
        finalSwing = [0, 0]
        for i in range(0, insideCount):
            [resultSet, bindSet] = getRightSet(sectionArray, file)
            finalSet = bindSet - resultSet
            if len(list({rightNumber} & finalSet)) != 0:
                insideResult["final_set"] += 1
            if len(list({rightNumber} & resultSet)) != 0:
                insideResult["right_set"] += 1
            if len(list({rightNumber} & bindSet)) != 0:
                insideResult["bind_set"] += 1
            allResult["right_set"].append(','.join(list(resultSet)))
            allResult["final_set"].append(','.join(list(finalSet)))
            allResult["bind_set"].append(','.join(list(bindSet)))
            #
            allLength["right_set"] += len(resultSet)
            allLength["final_set"] += len(finalSet)
            allLength["bind_set"] += len(bindSet)
            #
            if i == 0:
                finalSwing[0] = len(finalSet)
            else:
                if len(finalSet) < finalSwing[0]:
                    finalSwing[0] = len(finalSet)
                if len(finalSet) > finalSwing[1]:
                    finalSwing[1] = len(finalSet)
        totalResult.append(
            dict(insideResult=insideResult, allResult=allResult, allLength=allLength, finalSwing=finalSwing))

    for i in range(len(totalResult)):
        insideResult = totalResult[i]["insideResult"]
        allLength = totalResult[i]["allLength"]
        finalSwing = totalResult[i]["finalSwing"]
        infoInter = "交集命中概率(常规)：" + "{:.2%}".format(
            insideResult["right_set"] / insideCount) + "，交集平均数量：" + str(allLength["right_set"]/10)
        infoBind = "并集命中概率：" + \
            "{:.2%}".format(
                insideResult["bind_set"] / insideCount) + "，并集平均数量：" + str(allLength["bind_set"]/10)
        infoFinal = "差集命中概率：" + \
            "{:.2%}".format(
                insideResult["final_set"] / insideCount) + "，差集平均数量：" + str(allLength["final_set"]/10)+",振幅："+str(finalSwing[1]-finalSwing[0])
        ui.textBrowser.append("第"+str(i+1)+"次区间结果：")
        ui.textBrowser.append(infoInter)
        ui.textBrowser.append(infoBind)
        ui.textBrowser.append(infoFinal)

    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    for i in range(0, len(totalResult)):
        allResult = totalResult[i]["allResult"]
        for j in range(0, len(allResult["right_set"])):
            exportTxt3(filePath, '第'+str(i+1)+'区间交集码' +
                       str(j+1)+'_', allResult["right_set"][j])
        for j in range(0, len(allResult["bind_set"])):
            exportTxt3(filePath, '第'+str(i+1)+'区间并集码' +
                       str(j+1)+'_', allResult["bind_set"][j])
        for j in range(0, len(allResult["final_set"])):
            exportTxt3(filePath, '第'+str(i+1)+'区间差集码' +
                       str(j+1)+'_', allResult["final_set"][j])

    os.startfile(filePath)


def getRightSet(sectionArray, file):
    # 交
    resultSet = set()
    # 并
    bindSet = set()
    #
    setList = []
    for i in range(0, handlerCount):
        setList.append(main(sectionArray, file))

    for i in range(0, len(setList)):
        if len(resultSet) == 0:
            resultSet = setList[i]
            bindSet = setList[i]
        else:
            resultSet = resultSet.intersection(setList[i])
            bindSet = bindSet | setList[i]
    return [resultSet, bindSet]


# 核心函数  返回结果集合
def main(sectionArray, file):
    #
    [memberSection, hundredSection, fiftySection] = sectionArray
    #
    wb = openpyxl.load_workbook(file)
    sheetsNames = wb.sheetnames
    ws = wb[sheetsNames[0]]
    allList = list(range(1, totalCount + 1))
    random.shuffle(allList)
    # 所有号码 集合
    numbers_set = excelHelper.readWbFromIndex(
        ws, allList, memberSection)
    if isSingle == 1:
        return numbers_set
    #
    if totalCount > 100:
        ui.textBrowser.append("进行数据随机抽取"+str(hundredCount)+"次100")
        hundredSets = []
        for i in range(0, int(hundredCount)):
            indexes = random.sample(range(1, totalCount + 1), 100)
            hundredSets.append(excelHelper.readWbFromIndex(
                ws, indexes, hundredSection))
    ui.textBrowser.append("进行数据随机抽取"+str(fiftyCount)+"次50")
    fiftySets = []
    for i in range(0, int(fiftyCount)):
        indexes = random.sample(range(1, totalCount + 1), 50)
        fiftySets.append(excelHelper.readWbFromIndex(
            ws, indexes, fiftySection))
    # intersection
    if totalCount > 100:
        for item in hundredSets:
            numbers_set = numbers_set.intersection(item)
    for item in fiftySets:
        numbers_set = numbers_set.intersection(item)

    return numbers_set


def clearTextBrowser():
    ui.textBrowser.clear()


#
def sourceCheck():
    # 选择批量文件
    root = Tk()
    helperVersion = ccbox('选择版本', '提示', ('4位置', '5位置'))
    #
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    sourceCount = int(enterbox("批量的每个文件多少数据?", '确认', "0"))
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    #
    key = 1
    for file in cur:
        wb = openpyxl.load_workbook(file)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        allList = list(range(1, sourceCount + 1))
        # 集合
        if helperVersion == True:
            excelHelper.readWbFrom4IndexSource(ws, allList, filePath, key)
        else:
            excelHelper.readWbFromIndexSource(ws, allList, filePath, key)
        key += 1
    #
    ui.textBrowser.append('计算完成')
    root.destroy()
    root.mainloop()
    os.startfile(filePath)
    pass


def addSourceCheck():
    root = Tk()
    helperVersion = ccbox('选择版本', '提示', ('4位置', '5位置'))
    ui.textBrowser.append('计算&产生数据中......')
    #
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        root.destroy()
        ui.textBrowser.append('取消')
        return
    try:
        sourceCount = int(enterbox("选择文件一共多少数据?", '确认', "0"))
    except TypeError:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    else:
        pass
        #
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    # 排列组合 Cn2 后进行相加
    index = 1
    for i, j in combinations(cur, 2):
        wb = openpyxl.load_workbook(i)
        sheetsNames = wb.sheetnames
        ws_i = wb[sheetsNames[0]]
        wb1 = openpyxl.load_workbook(j)
        sheetsNames1 = wb1.sheetnames
        ws_j = wb1[sheetsNames1[0]]
        # 集合
        if helperVersion == True:
            excelHelper.readWbFrom4IndexAddSource(
                ws_i, ws_j, sourceCount, filePath, index)
        else:
            excelHelper.readWbFromIndexAddSource(
                ws_i, ws_j, sourceCount, filePath, index)
        index += 1
    # return result
    root.destroy()
    root.mainloop()
    ui.textBrowser.append(
        '计算完成')
    os.startfile(filePath)
    #
    pass


#  批量文件夹同序叠加
def mutilDirAdd():
    ui.textBrowser.append('计算&产生数据中......')
    helperVersion = ccbox('选择版本', '提示', ('4位置', '5位置'))
    #
    root = Tk()
    rootDir = filedialog.askdirectory()
    dirFileList = []
    for _root, dirs, files in os.walk(rootDir):
        if len(files) > 0:
            fileList = []
            for file in files:
                fileList.append(os.path.join(_root, file))
                pass
            dirFileList.append(fileList)
        pass
    dirCount = len(dirFileList)
    fileCount = len(dirFileList[0])
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    eachFileDataCount = enterbox("每个文件多少数据?", '确认', "0")
    if eachFileDataCount == None:
        ui.textBrowser.append('取消')
        return
    eachFileDataCount = int(eachFileDataCount)
    #
    index = 1
    for i in range(0, fileCount):
        wsList = []
        for j in range(0, dirCount):
            wb = openpyxl.load_workbook(dirFileList[j][i])
            sheetsNames = wb.sheetnames
            wsList.append(wb[sheetsNames[0]])
        #
        if helperVersion == True:
            excelHelper.readWbFromIndexAddDir(
                wsList, eachFileDataCount, filePath, index, 4)
        else:
            excelHelper.readWbFromIndexAddDir(
                wsList, eachFileDataCount, filePath, index, 5)
        index += 1

    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    pass


# 批量累加指定数量
def mutilAdd():
    ui.textBrowser.append('计算&产生数据中......')
    helperVersion = ccbox('选择版本', '提示', ('4位置', '5位置'))
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    #
    eachFileDataCount = int(enterbox("每个文件多少数据?", '确认', "0"))
    eachFileCount = int(enterbox("随机多少个文件为一组进行累加?", '确认', "0"))
    totalFileCount = int(enterbox("需要得到多少个累加文件？", '确认', "0"))
    allFileCount = len(cur)
    indexList = []
    for i in range(0, totalFileCount):
        indexList.append(random.sample(range(0, allFileCount), eachFileCount))

    #
    index = 1
    for i in range(0, len(indexList)):
        item = indexList[i]
        wsList = []
        for j in range(0, len(item)):
            wb = openpyxl.load_workbook(cur[item[j]])
            sheetsNames = wb.sheetnames
            wsList.append(wb[sheetsNames[0]])
        #
        if helperVersion == True:
            excelHelper.readWbFromIndex4Add(
                wsList, eachFileDataCount, filePath, index)
        else:
            excelHelper.readWbFromIndexAdd(
                wsList, eachFileDataCount, filePath, index)
        index += 1
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    #
    pass


# 批量excel set 区间处理


# 批量区间
def excelSet():
    helperVersion = ccbox('选择版本', '提示', ('4位置', '5位置'))
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    section = list(map(int, configuration['excelSection'].split(',')))
    excelCount = int(configuration['excelCount'])
    ui.textBrowser.append('计算数据中......')
    #
    root = Tk()
    # cur = diropenbox('数据目录')
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        root.destroy()
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    index = 1
    for path in cur:
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, excelCount+1)
        #
        if helperVersion == True:
            resultSet = excelHelper.readWbFrom4Index(
                ws, indexes, section)
        else:
            resultSet = excelHelper.readWbFromIndex(
                ws, indexes, section)
        file = open(filePath+"\\批量结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        index += 1
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    #
    pass


# 批量分组然后单区间
def partsSingleExcel():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    section = list(map(int, configuration['excelSection'].split(',')))
    ui.textBrowser.append('计算数据中......')
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        root.destroy()
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    #
    try:
        eachFileDataCount = int(enterbox("每个文件多少数据?", '确认', ""))
        eachGroupFileCount = int(enterbox("多少数据为一组?", '确认', ""))
    except TypeError:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    else:
        pass
        #
    # 进行分组
    pathListArray = []
    for i in range(0, len(cur), eachGroupFileCount):
        b = cur[i:i+eachGroupFileCount]
        if len(b) == eachGroupFileCount:
            pathListArray.append(b)

    excelHelper.partsReadWbFromIndex(
        pathListArray, eachFileDataCount, section, filePath)

    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    pass


# 批量双区间
def TwosideExcelSet():
    helperVersion = ccbox('选择版本', '提示', ('4位置', '5位置'))
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    leftSection = list(map(int, configuration['leftSection'].split(',')))
    rightSection = list(map(int, configuration['rightSection'].split(',')))
    excelCount = int(configuration['excelCount'])
    ui.textBrowser.append('批量双区间：计算数据中......')
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    # 排列组合 Cn2 后进行相加
    index = 1
    for path in cur:
        # left
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, excelCount+1)
        if helperVersion == True:
            resultSet = excelHelper.readWbFrom4Index(
                ws, indexes, leftSection)
        else:
            resultSet = excelHelper.readWbFromIndex(
                ws, indexes, leftSection)
        file = open(filePath+"\\批量左区间结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        # right
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, excelCount+1)
        if helperVersion == True:
            resultSet = excelHelper.readWbFrom4Index(
                ws, indexes, rightSection)
        else:
            resultSet = excelHelper.readWbFromIndex(
                ws, indexes, rightSection)
        file = open(filePath+"\\批量右区间结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        #
        index += 1
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)

    root.destroy()
    root.mainloop()
    pass


# 原始数据随机组合
def originPaste():
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        root.destroy()
        ui.textBrowser.append('取消')
        return
    try:
        totalCount = int(enterbox("选择文件每个一共多少数据?", '确认', "0"))
        dataCount = int(enterbox("多少数据为一个新文件?", '确认', "0"))
        fileCount = int(enterbox("每个文件需要生成多少文件？", '确认', "0"))
    except TypeError:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    else:
        pass
        #
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        root.destroy()
        ui.textBrowser.append('取消')
        return
        #
    #
    ui.textBrowser.append('开始计算....')
    #
    index = 1
    ResultList = []
    for path in cur:
        pathList = excelHelper.mutilOriginPaste(
            path, filePath, totalCount, fileCount, dataCount, index)
        index += 1
        ResultList.append(pathList)
        pass
    # 操作ResultList 进行叠加
    if len(ResultList) != 1:
        directoryExist = os.path.exists(filePath+'\\同序叠加结果')
        if directoryExist == False:
            os.makedirs(filePath+'\\同序叠加结果')
        _index = 1
        for n in range(0, fileCount):
            wsList = []
            for j in range(0, len(cur)):
                wb = openpyxl.load_workbook(ResultList[j][n])
                sheetsNames = wb.sheetnames
                wsList.append(wb[sheetsNames[0]])
            excelHelper.readWbFromIndex4Add(
                wsList, dataCount, filePath+'\\同序叠加结果', _index)
            _index += 1
            pass
    pass
    #
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    #
    pass


def originPasteSingle():
   #
    root = Tk()
    cur = filedialog.askopenfilename(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        root.destroy()
        ui.textBrowser.append('取消')
        return
    try:
        totalCount = int(enterbox("选择文件每个一共多少数据?", '确认', "0"))
        dataCount = int(enterbox("多少数据为一个新文件?", '确认', "0"))
        fileCount = int(enterbox("每个文件需要生成多少文件？", '确认', "0"))
    except TypeError:
        root.destroy()
        ui.textBrowser.append('取消')
        return
    else:
        pass
        #
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        root.destroy()
        ui.textBrowser.append('取消')
        return
        #
    #
    ui.textBrowser.append('开始计算....')
    #
    excelHelper.mutilOriginPaste(
        cur, filePath, totalCount, fileCount, dataCount, 0)

    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    #
    pass


def staticsPosition():
    ui.textBrowser.append('计算数据中......')
    #
    root = Tk()
    path = filedialog.askopenfilename(filetypes=[('xlsx files', '.xlsx')])
    if path == '':
        root.destroy()
        ui.textBrowser.append('取消')
        return
    try:
        excelCount = int(enterbox("选择文件的数据量", '确认', "0"))
    except TypeError:
        root.destroy()
        return False
    else:
        pass
    #
    wb = openpyxl.load_workbook(path)
    sheetsNames = wb.sheetnames
    ws = wb[sheetsNames[0]]
    indexes = range(1, excelCount+1)
    #
    aList = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    bList = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    cList = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    dList = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    eList = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    for i in range(0, len(indexes)):
        rx = indexes[i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        #
        w6 = str(ws.cell(row=rx, column=6).value)
        if w2 == "None":
            break
        #
        for i in range(0, len(w2)):
            aList[int(w2[i])] += 1
            pass
        for j in range(0, len(w3)):
            bList[int(w3[j])] += 1
            pass
        for m in range(0, len(w4)):
            cList[int(w4[m])] += 1
            pass
        for n in range(0, len(w5)):
            dList[int(w5[n])] += 1
            pass
        for n in range(0, len(w6)):
            eList[int(w6[n])] += 1
            pass
        pass

    strList = []
    for i in range(0, 10):
        strList.append('数字 '+str(i)+' 出现次数')

    aSet = dict(zip(strList, aList))
    bSet = dict(zip(strList, bList))
    cSet = dict(zip(strList, cList))
    dSet = dict(zip(strList, dList))
    eSet = dict(zip(strList, eList))
    #
    aRes = sorted(aSet.items(), key=lambda item: item[1], reverse=True)
    bRes = sorted(bSet.items(), key=lambda item: item[1], reverse=True)
    cRes = sorted(cSet.items(), key=lambda item: item[1], reverse=True)
    dRes = sorted(dSet.items(), key=lambda item: item[1], reverse=True)
    eRes = sorted(eSet.items(), key=lambda item: item[1], reverse=True)

    res = [aRes, bRes, cRes, dRes, eRes]

    for i in range(0, 5):
        ui.textBrowser.append('第'+str(i+1)+'位置结果：')
        ui.textBrowser.append('----')
        for j in range(0, len(res[i])):
            ui.textBrowser.append(res[i][j][0]+':'+str(res[i][j][1]))
        ui.textBrowser.append('\n')
    #
    root.destroy()
    root.mainloop()

    pass


# 批量命中excel统计
def excelMissing():
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        root.destroy()
        root.mainloop()
        ui.textBrowser.append('取消')
        return
    eachDataCount = int(enterbox("每个文件多少条数据?", '确认', "0"))
    rightNumber = enterbox("输入命中号码?", '确认', "0")
    if rightNumber == None:
        root.destroy()
        root.mainloop()
        ui.textBrowser.append('取消')
        return
    maxCount = 0
    minCount = 0
    totalCount = 0
    for path in cur:
        # left
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, eachDataCount+1)
        count = excelHelper.missingEachWbFromIndex(
            ws, indexes, rightNumber)
        totalCount += count
        if count > maxCount:
            maxCount = count
        if count < minCount and minCount != 0:
            minCount = count
        if minCount == 0:
            minCount = count
    avgCount = (totalCount)/len(cur)
    ui.textBrowser.append('计算完成')
    ui.textBrowser.append('最大命中结果：'+str(maxCount))
    ui.textBrowser.append('最小命中结果：'+str(minCount))
    ui.textBrowser.append('平均值：'+str(avgCount))
    ui.textBrowser.append('======================================')
    root.destroy()
    root.mainloop()
    pass


if __name__ == '__main__':    
    times = 0
    total = []
    _config = configparser.ConfigParser()
    _config.read('./pwd.ini')
    #
    nowDate = time.strftime("%Y%m%d", time.localtime())
    if _config['DEFAULT'].get('SECRET') == None or _config['DEFAULT']['SECRET'] != nowDate+'lin':
        passWord = passwordbox("请输入启动密码", '确认', "")
        if passWord != nowDate+'lin' and passWord != 'uercal':
            exit()            

    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    totalCount = int(configuration['totalCount'])
    hundredCount = int(configuration['hundredCount'])
    fiftyCount = int(configuration['fiftyCount'])
    handlerCount = int(configuration['handlerCount'])
    insideCount = int(configuration['insideCount'])
    isSingle = configuration['isSingle']
    isRange = configuration['isRange']
    qianRange = configuration['qianRange']
    baiRange = configuration['baiRange']
    shiRange = configuration['shiRange']
    geRange = configuration['geRange']
    setArray = configuration['setArray']
    #
    app = QApplication(sys.argv)
    MainWindow = QMainWindow()
    ui = window.Ui_MainWindow()
    ui.setupUi(MainWindow)
    #
    dialogWindow = QDialog()
    configui = config.Ui_config()
    configui.setupUi(dialogWindow)

    # 绑定dialog确认事件
    configui.buttonOk.accepted.connect(confirmConfig)
    configui.addSection.clicked.connect(addSection)
    configui.delSection.clicked.connect(delSection)
    MainWindow.show()
    # txt action绑定
    ui.actionjiaoji.triggered.connect(jiaoji)
    ui.actionchaji.triggered.connect(chaji)
    ui.actionbingji.triggered.connect(bingji)
    ui.actioncheck.triggered.connect(check)
    ui.actionBind.triggered.connect(mutilBind)
    ui.actionmutilCha.triggered.connect(mutilCha)
    ui.actionrandomMutilTxt.triggered.connect(randomMutilTxt)
    ui.actionmultiPositionHandler.triggered.connect(multiPositionHandler)
    ui.actionfindEmpty.triggered.connect(findEmptyHandler)
    ui.actionsingleNumberRate.triggered.connect(singleNumberRate)
    ui.actiondanshuangPosition.triggered.connect(danshuangPosition)
    ui.actionpositionCountOrder.triggered.connect(actionpositionCountOrder)
    ui.actionorderFileCombine.triggered.connect(actionorderFileCombine)
    ui.actionfiveEmptyBindInter.triggered.connect(fiveEmptyBindInter)
    ui.actionshaCodeFullCha.triggered.connect(shaCodeFullCha)

    # 逆序算法
    ui.actionreverseResult.triggered.connect(reverseCheck)

    # ------------------
    # 随机求交
    ui.actiontxtRandom.triggered.connect(txtRandomTimes)
    ui.actiontxtRandomGroup.triggered.connect(txtRandomTimesGroup)
    # 随机求并
    ui.actionrandomPickBind.triggered.connect(txtRandomPickBind)
    ui.actionrandomGroupBind.triggered.connect(txtRandomGroupBind)
    # 随机抽
    ui.actiondiyPIckGroupInter.triggered.connect(txtDiyPickGroupInter)

    ui.clearButton.clicked.connect(clearTextBrowser)
    # 爬虫类绑定
    ui.actionloadingRecent.triggered.connect(loadRecent)
    ui.actionexportPast.triggered.connect(climb)
    # excel类绑定
    ui.actionconfig.triggered.connect(showConfig)
    ui.actionmainCheck.triggered.connect(mainCheck)
    # 交叉
    ui.actionSource.triggered.connect(sourceCheck)

    # --------
    # 批量累加
    ui.actionmutilAdd.triggered.connect(mutilAdd)
    # 累加
    ui.actionAddSource.triggered.connect(addSourceCheck)
    # 批量文件夹同序叠加
    ui.actionMutilDirAdd.triggered.connect(mutilDirAdd)

    # ------
    # 批量单区间
    ui.actionExcelSet.triggered.connect(excelSet)
    # 批量分组单区间
    ui.actionpartsSingleExcel.triggered.connect(partsSingleExcel)

    # ----
    # 批量双区间
    ui.actionactionTwoside.triggered.connect(TwosideExcelSet)
    # 原始数据随机组合
    ui.actionsingleOriginPaste.triggered.connect(originPasteSingle)
    ui.actionmutilOriginPaste.triggered.connect(originPaste)
    # ui.actiondirectoryAdd.triggered.connect(direcotryOriginAdd)
    # 统计各位置数字
    ui.actionStaticsPosition.triggered.connect(staticsPosition)
    # 批量命中
    ui.actionexcelMissing.triggered.connect(excelMissing)

    sys.exit(app.exec_())
