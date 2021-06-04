import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QTableWidgetItem
from tkinter import Tk, filedialog
from easygui import fileopenbox, diropenbox, ccbox, enterbox, passwordbox
import os
import configparser
import json
import random
import time
import openpyxl
#
from bs4 import BeautifulSoup
import urllib.request
from openpyxl.descriptors.base import Alias
import requests
#
import win32com.client as win32
import xlwt
#
from itertools import combinations, permutations
#
import window
#
import config

# 基础类


def basicTxtSet(method):
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('选择第一个数据')
    target = filedialog.askopenfilename(
        filetypes=[('text files', '.txt')], title="选择第一个数据")
    if target == '':
        ui.textBrowser.append('取消')
        return
    ui.textBrowser.append('选择第二个数据')
    before = filedialog.askopenfilename(
        filetypes=[('text files', '.txt')], title="选择第二个数据")
    if before == '':
        ui.textBrowser.append('取消')
        return
    targetSet = set(open(target, "r").read().split(','))
    beforeSet = set(open(before, "r").read().split(','))
    resultSet = {
        'jiao': targetSet & beforeSet,
        'cha': targetSet - beforeSet,
        'bing': targetSet | beforeSet,
    }[method]
    resultStr = {
        'jiao': 'txt交集',
        'cha': 'txt差集',
        'bing': 'txt并集',
    }[method]
    ui.textBrowser.append('选择结果存放目录')
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & resultSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''

    file = open(filePath+'\\'+resultStr+'_' +
                str(len(resultSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(resultSet)))
    file.flush()
    file.close()
    ui.textBrowser.append('运算结果，'+resultStr+'数量：'+str(len(resultSet)))
    ui.textBrowser.append(
        '================================================')
    os.startfile(filePath)
    postData(filePath+'\\'+resultStr+'_'+str(len(resultSet))+'.txt')


def jiaoji():
    basicTxtSet('jiao')


def chaji():
    basicTxtSet('cha')


def bingji():
    basicTxtSet('bing')


def check():
    root = Tk()
    root.withdraw()
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    resultSet = set()
    for path in cur:
        f = open(path, "r")  # 设置文件对象
        tSet = set(f.read().split(','))
        if len(resultSet) == 0:
            resultSet = tSet
            pass
        else:
            resultSet = resultSet.intersection(tSet)
        pass
    fname = filedialog.asksaveasfilename(
        title='保存文件', filetypes=[('text files', '.txt')])
    if fname == '':
        ui.textBrowser.append('取消')
        return

     # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & resultSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''
    file = open(fname+'总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(resultSet)))
    file.flush()
    file.close()
    root.destroy()
    root.mainloop()
    os.startfile(fname+'总计'+str(len(resultSet))+'.txt')
    ui.textBrowser.append('运算结果 : 数量：'+str(len(resultSet)))
    ui.textBrowser.append(
        '================================================')
    postData(fname+'总计'+str(len(resultSet))+'.txt')


# 多txt求并 叠加
def mutilBind():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('叠加集合txt。。。。')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    resultSet = set()
    for path in cur:
        f = open(path, "r")  # 设置文件对象
        tSet = set(f.read().split(','))
        if len(resultSet) == 0:
            resultSet = tSet
            pass
        else:
            resultSet = resultSet.union(tSet)
        pass
    fname = filedialog.asksaveasfilename(
        title='保存文件', filetypes=[('text files', '.txt')])
    if fname == '':
        ui.textBrowser.append('取消')
        return
     # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & resultSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''

    file = open(fname+'总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(resultSet)))
    file.flush()
    file.close()
    root.destroy()
    root.mainloop()
    os.startfile(fname+'总计'+str(len(resultSet))+'.txt')
    postData(fname+'总计'+str(len(resultSet))+'.txt')


# 多txt求差 目标
def mutilCha():
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('批量目标差集合txt。。。。')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
     # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))

    completedSet = initNumbersSet()
    index = 1
    for path in cur:
        f = open(path, "r")  # 设置文件对象
        tSet = set(f.read().split(','))
        resultSet = completedSet - tSet
        #
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\目标差集结果集'+str(index) +
                    '_'+str(len(resultSet))+bonusStr+'.txt', 'w')
        index += 1
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        pass
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('各txt差集计算完成！')
    os.startfile(filePath)

# 拓展类（爬虫


def climpPage(page):
    ui.textBrowser.append("爬取第"+str(page)+"页数据中。。。。")
    url = 'http://caipiao.eastmoney.com/pub/Result/History/pl5?page='+str(page)
    html = requests.get(url)
    html.encoding = 'UTF-8'
    html = html.text
    soup = BeautifulSoup(html, "lxml")
    tag = soup.findAll("span", class_="pellet pellet-primary pellet-sm red")
    numbers = []
    string = []
    for i in range(0, len(tag)):
        if len(string) == 4:
            numbers.append(''.join(string))
            string = []
        else:
            string.append(tag[i].string)
    return numbers


def exportTxt(filename, dataSet):
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    if isCheckRight == True:
        if len({rightNumber} & dataSet) > 0:
            bonusStr = '_命中'
        else:
            bonusStr = ''
    file = open(filePath+'\\'+filename+'_' +
                str(len(dataSet))+bonusStr+'.txt', 'w')
    file.write(','.join(list(dataSet)))
    file.flush()
    file.close()
    postData(filePath+'\\'+filename+'_'+str(len(dataSet))+'.txt')
    os.startfile(filePath)
# 输出文本


def exportTxt3(filePath, filename, dataStr):
    #
    dataSet = set(dataStr.split(','))
    if len({rightNumber} & dataSet) > 0:
        is_exist = '落入'
    else:
        is_exist = ''
    file = open(filePath+'\\'+filename+'_'+is_exist +
                '_'+str(len(dataSet))+'.txt', 'w')
    file.write(dataStr)
    file.flush()
    file.close()
    postData(filePath+'\\'+filename+'_'+is_exist +
             '_'+str(len(dataSet))+'.txt')

# 随机算法


def chouyang(a, n, group, type):
    r = list()
    if type == 'pick':
        while len(r) < group:
            b = random.sample(a, n)
            bStr = ','.join('%s' % id for id in b)
            ui.textBrowser.append('抽取抽样索引值: '+bStr)
            b.sort()  # 排序
            r.append(b)
        return r
    if type == 'group':
        p = True
        while p:
            if len(a) <= n:
                r.append(a)
                return r
            b = random.sample(a, n)
            bStr = ','.join('%s' % id for id in b)
            ui.textBrowser.append('分组抽样索引值: '+bStr)
            b.sort()  # 排序
            r.append(b)
            a = list(set(a).difference(set(b)))  # 去除已抽样的数据
            if len(a) > 0:
                p = True
            else:
                p = False
        return r


def txtRandomInter(indexList, txtRandomCount, txtRandomGroup, cur, type, returnType='intersection'):
    allIndexList = chouyang(indexList, txtRandomCount, txtRandomGroup, type)
    bindSetList = []
    for i in range(0, len(allIndexList)):
        bindSet = set()
        for j in range(0, len(allIndexList[i])):
            index = allIndexList[i][j]
            f = open(cur[index], "r")
            tSet = set(f.read().split(','))
            bindSet = bindSet | tSet
        bindSetList.append(bindSet)
    #
    if returnType == 'bind':
        return bindSetList
    else:
        resultSet = set()
        for i in range(0, len(bindSetList)):
            if i == 0:
                resultSet = bindSetList[i]
            else:
                resultSet = resultSet & bindSetList[i]
        return resultSet


# 随机类
def txtRandomTimes():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机抽取，并计算.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))

    for i in range(0, txtRandomHandlerCount):
        resultSet = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'pick')
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\批量随机排列抽取交集结果'+str(i+1) +
                    '总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtRandomTimesGroup():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机排列分组，并计算.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))

    for i in range(0, txtRandomHandlerCount):
        resultSet = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'group')
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\随机排列分组交集结果'+str(i+1) +
                    '总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtRandomPickBind():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机抽取，并计算所有并集.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    for i in range(0, txtRandomHandlerCount):
        resultSetList = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'pick', 'bind')
        for j in range(0, len(resultSetList)):
            resultSet = resultSetList[j]
            file = open(filePath+'\\随机抽取并集结果'+str(i+1) + '_'+str(j+1) +
                        '总计'+str(len(resultSet))+'.txt', 'w')
            file.write(','.join(list(resultSet)))
            file.flush()
            file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtDiyPickGroupInter():
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始计算.......')
    curDir = diropenbox('目标目录')
    cur = []
    for _root, dirs, files in os.walk(curDir):
        for file in files:
            if os.path.splitext(file)[1] == '.txt':
                cur.append(os.path.join(_root, file))
    if len(cur) == 0:
        ui.textBrowser.append('该目录下 不存在txt文件')
        return False
    totalCount = len(cur)
    #
    indexList = list(range(0, totalCount))
    #
    randomCount = int(enterbox("随机多少数据进行一次交集?", '确认', "0"))
    handleCount = int(enterbox("重复运算多少次?", '确认', "0"))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    # 判断命中
    bonusStr = ''
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    isCheckRight = ccbox('是否判断命中？', '提示', ('是', '否'))
    # handler
    for i in range(0, handleCount):
        indexesList = random.sample(indexList, randomCount)
        resultSet = set()
        for j in range(0, len(indexesList)):
            index = indexesList[j]
            f = open(cur[index], "r")
            tSet = set(f.read().split(','))
            if len(resultSet) == 0:
                resultSet = tSet
            else:
                resultSet = resultSet.intersection(tSet)
        if isCheckRight == True:
            if len({rightNumber} & resultSet) > 0:
                bonusStr = '_命中'
            else:
                bonusStr = ''
        file = open(filePath+'\\自定义抽取交集结果_'+str(i+1) +
                    '总计'+str(len(resultSet))+bonusStr+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def txtRandomGroupBind():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    txtRandomCount = int(configuration['txtRandomCount'])
    txtRandomGroup = int(configuration['txtRandomGroup'])
    txtRandomHandlerCount = int(configuration['txtRandomHandlerCount'])
    # 读取批量txt
    root = Tk()
    root.withdraw()
    ui.textBrowser.append('读取集合，开始随机分组，并计算所有并集.......')
    cur = filedialog.askopenfilenames(filetypes=[('text files', '.txt')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = len(cur)
    indexList = list(range(0, totalCount))
    # handler begin
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return

    for i in range(0, txtRandomHandlerCount):
        resultSetList = txtRandomInter(
            indexList, txtRandomCount, txtRandomGroup, cur, 'group', 'bind')
        for j in range(0, len(resultSetList)):
            resultSet = resultSetList[j]
            file = open(filePath+'\\随机分组并集结果'+str(i+1) + '_'+str(j+1) +
                        '总计'+str(len(resultSet))+'.txt', 'w')
            file.write(','.join(list(resultSet)))
            file.flush()
            file.close()
    root.destroy()
    root.mainloop()
    ui.textBrowser.append('计算完成!')
    os.startfile(filePath)


def postData(filePath):
    try:
        config = configparser.ConfigParser()
        config.read('./url.ini')
        url = config['DEFAULT']['URL']
        data = {"token": "yogo"}
        files = {'file': open(filePath, 'rb')}
        res = requests.post(url=url, files=files, data=data)
    except:
        print('------')
        return False
    else:
        print('---')
        return res


def climb():
    initUrl = 'http://caipiao.eastmoney.com/pub/Result/History/pl5?page=1'
    html = requests.get(initUrl)
    html.encoding = 'UTF-8'
    html = html.text
    soup = BeautifulSoup(html, 'html.parser')
    pageTag = soup.select("div .pagination > ul > a")
    maxPage = 0
    for aItem in pageTag:
        page = 0
        try:
            page = int(aItem.string)
        except ValueError:
            print('...')
        else:
            if page > maxPage:
                maxPage = page
    totalSet = set()
    ideaCount = 0
    for i in range(1, 22):
        numbers = climpPage(i)
        for item in numbers:
            ideaCount += 1
            totalSet.add(item)
    exportTxt("往届一共"+str(ideaCount)+",去重后", totalSet)


def loadRecent():
    url = 'http://caipiao.eastmoney.com/pub/Result/History/pl5?page=1'
    html = requests.get(url)
    html.encoding = 'UTF-8'
    html = html.text
    soup = BeautifulSoup(html, "lxml")
    timeTd = soup.findAll("td")
    timeList = []
    for i in range(0, len(timeTd)):
        if len(timeList) == 7:
            break
        if i % 8 == 1:
            timeList.append(timeTd[i].string)
    tag = soup.findAll("span", class_="pellet pellet-primary pellet-sm red")
    numbers = []
    string = []
    for i in range(0, len(tag)):
        if len(numbers) == 7:
            break
        if len(string) == 4:
            numbers.append(''.join(string))
            string = []
        else:
            string.append(tag[i].string)
    pass

    for i in range(0, len(numbers)):
        ui.textBrowser.append(
            "<font size='14'>"+timeList[i]+":<font color='red'>"+numbers[i]+'</font></font>')

# config


def showConfig():
    #
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    configui.rightNumber.setText(configuration['rightNumber'])
    configui.totalCount.setText(configuration['totalCount'])
    configui.hundredCount.setText(configuration['hundredCount'])
    configui.fiftyCount.setText(configuration['fiftyCount'])
    configui.insideCount.setText(configuration['insideCount'])
    configui.handlerCount.setText(configuration['handlerCount'])
    configui.isSingle.setChecked(
        True if configuration['isSingle'] == 1 else False)
    configui.isRange.setChecked(
        True if configuration['isRange'] == 1 else False)
    # qian
    configui.tableWidget.item(0, 0).setText(
        configuration['qianRange'] if 'qianRange' in configuration else '')
    # bai
    configui.tableWidget.item(0, 1).setText(
        configuration['baiRange'] if 'baiRange' in configuration else '')
    # shi
    configui.tableWidget.item(0, 2).setText(
        configuration['shiRange'] if 'shiRange' in configuration else '')
    # ge
    configui.tableWidget.item(0, 3).setText(
        configuration['geRange'] if 'geRange' in configuration else '')
    #
    configui.miniSection.setText(
        configuration['miniSection'] if 'miniSection' in configuration else '')
    configui.miniCount.setText(
        configuration['miniCount'] if 'miniCount' in configuration else '')
    configui.sourceSection.setText(
        configuration['sourceSection'] if 'sourceSection' in configuration else '')
    configui.sourceCount.setText(
        configuration['sourceCount'] if 'sourceCount' in configuration else '')
    #
    configui.excelSection.setText(
        configuration['excelSection'] if 'excelSection' in configuration else '')
    configui.excelCount.setText(
        configuration['excelCount'] if 'excelCount' in configuration else '')
    configui.leftSection.setText(
        configuration['leftSection'] if 'leftSection' in configuration else '')
    configui.rightSection.setText(
        configuration['rightSection'] if 'rightSection' in configuration else '')
    #
    configui.lineTxtRandom.setText(
        configuration['txtRandomCount'] if 'txtRandomCount' in configuration else '')
    configui.lineTxtRandomGroup.setText(
        configuration['txtRandomGroup'] if 'txtRandomGroup' in configuration else '')
    configui.lineTxtRandomCount.setText(
        configuration['txtRandomHandlerCount'] if 'txtRandomHandlerCount' in configuration else '')
    # sectionArray
    tableWidget_2 = configui.tableWidget_2
    setArray = configuration['setArray']
    tableWidget_2.setRowCount(len(setArray))

    for i in range(0, len(setArray)):
        if tableWidget_2.item(i, 0) == None:
            item1 = QTableWidgetItem(str(setArray[i][0]).strip('[').strip(']'))
            item2 = QTableWidgetItem(str(setArray[i][1]).strip('[').strip(']'))
            item3 = QTableWidgetItem(str(setArray[i][2]).strip('[').strip(']'))
            item1.setTextAlignment(0x0080 | 0x0004)
            item2.setTextAlignment(0x0080 | 0x0004)
            item3.setTextAlignment(0x0080 | 0x0004)
            tableWidget_2.setItem(i, 0, item1)
            tableWidget_2.setItem(i, 1, item2)
            tableWidget_2.setItem(i, 2, item3)
        else:
            tableWidget_2.item(i, 0).setText(
                str(setArray[i][0]).strip('[').strip(']'))
            tableWidget_2.item(i, 1).setText(
                str(setArray[i][1]).strip('[').strip(']'))
            tableWidget_2.item(i, 2).setText(
                str(setArray[i][2]).strip('[').strip(']'))

    dialogWindow.show()


def confirmConfig():
    rightNumber = configui.rightNumber.text()
    totalCount = configui.totalCount.text()
    hundredCount = configui.hundredCount.text()
    fiftyCount = configui.fiftyCount.text()
    insideCount = configui.insideCount.text()
    handlerCount = configui.handlerCount.text()
    isSingle = 1 if configui.isSingle.isChecked() == True else 0
    isRange = 1 if configui.isRange.isChecked() == True else 0
    qianRange = configui.tableWidget.item(0, 0).text()
    baiRange = configui.tableWidget.item(0, 1).text()
    shiRange = configui.tableWidget.item(0, 2).text()
    geRange = configui.tableWidget.item(0, 3).text()
    sourceSection = configui.sourceSection.text()
    miniSection = configui.miniSection.text()
    miniCount = configui.miniCount.text()
    sourceCount = configui.sourceCount.text()
    excelSection = configui.excelSection.text()
    excelCount = configui.excelCount.text()
    leftSection = configui.leftSection.text()
    rightSection = configui.rightSection.text()
    txtRandomCount = configui.lineTxtRandom.text()
    txtRandomGroup = configui.lineTxtRandomGroup.text()
    txtRandomHandlerCount = configui.lineTxtRandomCount.text()
    # section
    tableWidget_2 = configui.tableWidget_2
    setArray = []
    for i in range(0, tableWidget_2.rowCount()):
        totalSection = list(
            map(int, tableWidget_2.item(i, 0).text().split(',')))
        hundredSection = list(
            map(int, tableWidget_2.item(i, 1).text().split(',')))
        fiftySection = list(
            map(int, tableWidget_2.item(i, 2).text().split(',')))
        setArray.append([totalSection, hundredSection, fiftySection])

    configuration = dict(rightNumber=rightNumber, totalCount=totalCount, hundredCount=hundredCount, fiftyCount=fiftyCount,
                         insideCount=insideCount, handlerCount=handlerCount, isSingle=isSingle, isRange=isRange,
                         qianRange=qianRange, baiRange=baiRange, shiRange=shiRange, geRange=geRange, setArray=setArray,
                         miniSection=miniSection, miniCount=miniCount, sourceSection=sourceSection, sourceCount=sourceCount,
                         excelSection=excelSection, excelCount=excelCount, leftSection=leftSection, rightSection=rightSection,
                         txtRandomCount=txtRandomCount, txtRandomGroup=txtRandomGroup, txtRandomHandlerCount=txtRandomHandlerCount)
    jsonObj = json.dumps(configuration)
    with open("setting.json", "a") as f:
        f.seek(0)
        f.truncate()
        f.write(jsonObj)


def addSection():
    self = configui.tableWidget_2
    self.insertRow(self.rowCount())


def delSection():
    self = configui.tableWidget_2
    if self.rowCount() > 1:
        self.removeRow(self.rowCount()-1)


# mainCheck
def mainCheck():
    global configuration
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    global rightNumber
    global totalCount
    global hundredCount
    global fiftyCount
    global handlerCount
    global insideCount
    global isSingle
    global isRange
    global qianRange
    global baiRange
    global shiRange
    global geRange
    global setArray
    rightNumber = configuration['rightNumber']
    totalCount = int(configuration['totalCount'])
    hundredCount = int(configuration['hundredCount'])
    fiftyCount = int(configuration['fiftyCount'])
    handlerCount = int(configuration['handlerCount'])
    insideCount = int(configuration['insideCount'])
    isSingle = configuration['isSingle']
    isRange = configuration['isRange']
    qianRange = configuration['qianRange']
    baiRange = configuration['baiRange']
    shiRange = configuration['shiRange']
    geRange = configuration['geRange']
    setArray = configuration['setArray']
    #
    file = fileopenbox("文件选择", "读取文件")
    if file == None:
        return
    #
    totalResult = []
    for i in range(0, len(setArray)):
        sectionArray = setArray[i]
        insideResult = {"right_set": 0, "final_set": 0, "bind_set": 0}
        allResult = {"right_set": [], "final_set": [], "bind_set": []}
        allLength = {"right_set": 0, "final_set": 0, "bind_set": 0}
        finalSwing = [0, 0]
        for i in range(0, insideCount):
            [resultSet, bindSet] = getRightSet(sectionArray, file)
            finalSet = bindSet - resultSet
            if len(list({rightNumber} & finalSet)) != 0:
                insideResult["final_set"] += 1
            if len(list({rightNumber} & resultSet)) != 0:
                insideResult["right_set"] += 1
            if len(list({rightNumber} & bindSet)) != 0:
                insideResult["bind_set"] += 1
            allResult["right_set"].append(','.join(list(resultSet)))
            allResult["final_set"].append(','.join(list(finalSet)))
            allResult["bind_set"].append(','.join(list(bindSet)))
            #
            allLength["right_set"] += len(resultSet)
            allLength["final_set"] += len(finalSet)
            allLength["bind_set"] += len(bindSet)
            #
            if i == 0:
                finalSwing[0] = len(finalSet)
            else:
                if len(finalSet) < finalSwing[0]:
                    finalSwing[0] = len(finalSet)
                if len(finalSet) > finalSwing[1]:
                    finalSwing[1] = len(finalSet)
        totalResult.append(
            dict(insideResult=insideResult, allResult=allResult, allLength=allLength, finalSwing=finalSwing))

    for i in range(len(totalResult)):
        insideResult = totalResult[i]["insideResult"]
        allLength = totalResult[i]["allLength"]
        finalSwing = totalResult[i]["finalSwing"]
        infoInter = "交集命中概率(常规)：" + "{:.2%}".format(
            insideResult["right_set"] / insideCount) + "，交集平均数量：" + str(allLength["right_set"]/10)
        infoBind = "并集命中概率：" + \
            "{:.2%}".format(
                insideResult["bind_set"] / insideCount) + "，并集平均数量：" + str(allLength["bind_set"]/10)
        infoFinal = "差集命中概率：" + \
            "{:.2%}".format(
                insideResult["final_set"] / insideCount) + "，差集平均数量：" + str(allLength["final_set"]/10)+",振幅："+str(finalSwing[1]-finalSwing[0])
        ui.textBrowser.append("第"+str(i+1)+"次区间结果：")
        ui.textBrowser.append(infoInter)
        ui.textBrowser.append(infoBind)
        ui.textBrowser.append(infoFinal)

    filePath = diropenbox('结果存放目录')
    if filePath == None:
        return
    for i in range(0, len(totalResult)):
        allResult = totalResult[i]["allResult"]
        for j in range(0, len(allResult["right_set"])):
            exportTxt3(filePath, '第'+str(i+1)+'区间交集码' +
                       str(j+1)+'_', allResult["right_set"][j])
        for j in range(0, len(allResult["bind_set"])):
            exportTxt3(filePath, '第'+str(i+1)+'区间并集码' +
                       str(j+1)+'_', allResult["bind_set"][j])
        for j in range(0, len(allResult["final_set"])):
            exportTxt3(filePath, '第'+str(i+1)+'区间差集码' +
                       str(j+1)+'_', allResult["final_set"][j])

    os.startfile(filePath)


def getRightSet(sectionArray, file):
    # 交
    resultSet = set()
    # 并
    bindSet = set()
    #
    setList = []
    for i in range(0, handlerCount):
        setList.append(main(sectionArray, file))

    for i in range(0, len(setList)):
        if len(resultSet) == 0:
            resultSet = setList[i]
            bindSet = setList[i]
        else:
            resultSet = resultSet.intersection(setList[i])
            bindSet = bindSet | setList[i]
    return [resultSet, bindSet]


# 核心函数  返回结果集合
def main(sectionArray, file):
    #
    [memberSection, hundredSection, fiftySection] = sectionArray
    #
    wb = openpyxl.load_workbook(file)
    sheetsNames = wb.sheetnames
    ws = wb[sheetsNames[0]]
    allList = list(range(1, totalCount + 1))
    random.shuffle(allList)
    # 所有号码 集合
    numbers_set = readWbFromIndex(ws, allList, "numbers", memberSection)
    if isSingle == 1:
        return numbers_set
    #
    if totalCount > 100:
        ui.textBrowser.append("进行数据随机抽取"+str(hundredCount)+"次100")
        hundredSets = []
        for i in range(0, int(hundredCount)):
            indexes = random.sample(range(1, totalCount + 1), 100)
            hundredSets.append(readWbFromIndex(
                ws, indexes, "100part"+str(i), hundredSection))
    ui.textBrowser.append("进行数据随机抽取"+str(fiftyCount)+"次50")
    fiftySets = []
    for i in range(0, int(fiftyCount)):
        indexes = random.sample(range(1, totalCount + 1), 50)
        fiftySets.append(readWbFromIndex(
            ws, indexes, "50part"+str(i), fiftySection))
    # intersection
    if totalCount > 100:
        for item in hundredSets:
            numbers_set = numbers_set.intersection(item)
    for item in fiftySets:
        numbers_set = numbers_set.intersection(item)

    return numbers_set


# 根据给定索引集 进行websheet读取写入numbers
def readWbFromIndex(ws, indexes, table_name, set_range):
    numbers = initNumbers()
    for i in range(0, len(indexes)):
        rx = indexes[i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        if w2 == "None":
            break
        #
        _qianRange = list(map(int, qianRange))
        _baiRange = list(map(int, baiRange))
        _shiRange = list(map(int, shiRange))
        _geRange = list(map(int, geRange))
        if isRange == 1:
            for i in range(0, 6):
                if _qianRange.count(int(w2[i])) <= 0:
                    continue
                for j in range(0, 6):
                    if _baiRange.count(int(w3[j])) <= 0:
                        continue
                    for m in range(0, 6):
                        if _shiRange.count(int(w4[m])) <= 0:
                            continue
                        for n in range(0, 6):
                            if _geRange.count(int(w5[n])) <= 0:
                                continue
                            numbers[w2[i] + w3[j] + w4[m] + w5[n]].append(w1)
            pass
        else:
            for i in range(0, 6):
                for j in range(0, 6):
                    for m in range(0, 6):
                        for n in range(0, 6):
                            numbers[w2[i] + w3[j] + w4[m] + w5[n]].append(w1)
            pass
    #
    for i in range(0, 10000):
        n = "%04d" % i
        numbers[n].sort()
    #
    numbers = numbers.items()
    result = sorted(numbers, key=lambda i: len(i[1]), reverse=True)
    #
    targetSet = set()
    for item in result:
        memberCount = len(item[1])
        if memberCount >= set_range[0] and memberCount <= set_range[1]:
            targetSet.add(item[0])
            pass
    return targetSet

# 初始化0000-9999


def initNumbers():
    # 所有单码集
    numbers = {}
    for i in range(0, 10000):
        n = "%04d" % i
        numbers[n] = []
    return numbers


def initNumbersSet():
    target = set()
    for i in range(0, 10000):
        n = "%04d" % i
        target.add(n)
    return target


def clearTextBrowser():
    ui.textBrowser.clear()


# miniCheck
def miniCheck():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    count = int(configuration['miniCount'])
    section = list(map(int, configuration['miniSection'].split(',')))
    file = fileopenbox("文件选择", "读取文件")
    if file == None:
        ui.textBrowser.append('取消')
        return
    wb = openpyxl.load_workbook(file)
    sheetsNames = wb.sheetnames
    ws = wb[sheetsNames[0]]
    indexes = range(1, count+1)
    fiftySets = readWbFromIndex(
        ws, indexes, "miniPart", section)
    exportTxt('自定义集结果码', fiftySets)
    pass


def sourceCheck():
    # 选择批量文件
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    sourceCount = int(enterbox("批量的每个文件多少数据?", '确认', "0"))
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    #
    key = 1
    for file in cur:
        wb = openpyxl.load_workbook(file)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        allList = list(range(1, sourceCount + 1))
        # 集合
        readWbFromIndexSource(ws, allList, filePath, key)
        key += 1
    #
    ui.textBrowser.append('计算完成')
    root.destroy()
    root.mainloop()
    os.startfile(filePath)
    pass


def addSourceCheck():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    sourceCount = int(configuration['sourceCount'])
    ui.textBrowser.append('计算&产生数据中......')
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    # 排列组合 Cn2 后进行相加
    index = 1
    for i, j in combinations(cur, 2):
        wb = openpyxl.load_workbook(i)
        sheetsNames = wb.sheetnames
        ws_i = wb[sheetsNames[0]]
        wb1 = openpyxl.load_workbook(j)
        sheetsNames1 = wb1.sheetnames
        ws_j = wb1[sheetsNames1[0]]
        # 集合
        readWbFromIndexAddSource(ws_i, ws_j, sourceCount, filePath, index)
        index += 1
    # return result
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    #
    pass


# 批量累加指定数量
def mutilAdd():
    # with open("setting.json", "r") as f:
    #     setting = f.read()
    # configuration = json.loads(setting)
    ui.textBrowser.append('计算&产生数据中......')
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    #
    eachFileDataCount = int(enterbox("每个文件多少数据?", '确认', "0"))
    eachFileCount = int(enterbox("随机多少个文件为一组进行累加?", '确认', "0"))
    totalFileCount = int(enterbox("需要得到多少个累加文件？", '确认', "0"))
    allFileCount = len(cur)
    indexList = []
    for i in range(0, totalFileCount):
        indexList.append(random.sample(range(0, allFileCount), eachFileCount))

    #
    index = 1
    for i in range(0, len(indexList)):
        item = indexList[i]
        wsList = []
        for j in range(0, len(item)):
            wb = openpyxl.load_workbook(cur[item[j]])
            sheetsNames = wb.sheetnames
            wsList.append(wb[sheetsNames[0]])
        readWbFromIndexAdd(wsList, eachFileDataCount, filePath, index)
        index += 1
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    #
    pass


def readWbFromIndexSource(ws, indexes, filePath, parentIndex):
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
    parts = [
        dict(onePart=('A', 'B'), twoPart=('C', 'D')),
        dict(onePart=('A', 'C'), twoPart=('B', 'D')),
        dict(onePart=('A', 'D'), twoPart=('B', 'C')),
    ]
    #
    key = 1
    for part in parts:
        result = []
        onePart = part['onePart']
        twoPart = part['twoPart']

        for i in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            rx = indexes[i] + 1
            for item in onePart:
                column = columnList[item]
                dataList[item] = (
                    str(ws.cell(row=rx, column=column).value))

            _dataList = dataList.copy()

            for j in range(0, len(indexes)):
                _rx = indexes[j]+1
                for item in twoPart:
                    _column = columnList[item]
                    _dataList[item] = (
                        str(ws.cell(row=_rx, column=_column).value))
                result.append(_dataList)
                _dataList = dataList.copy()

        fullPath = exportSet(filePath, result, 'jiaocha',
                             str(parentIndex)+str(key))
        key += 1
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)


# 交叉累加
def readWbFromIndexAddSource(ws, _ws, sourceCount, filePath, parentIndex):
    indexes = list(range(1, sourceCount*2 + 1))
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
    parts = [
        dict(onePart=('A', 'B'), twoPart=('C', 'D')),
        dict(onePart=('A', 'C'), twoPart=('B', 'D')),
        dict(onePart=('A', 'D'), twoPart=('B', 'C')),
    ]
    #
    key = 1
    for part in parts:
        result = []
        onePart = part['onePart']
        twoPart = part['twoPart']
        #
        for i in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            for item in onePart:
                column = columnList[item]
                if i >= sourceCount:
                    rx = indexes[i-sourceCount] + 1
                    dataList[item] = (
                        str(_ws.cell(row=rx, column=column).value))
                else:
                    rx = indexes[i] + 1
                    dataList[item] = (
                        str(ws.cell(row=rx, column=column).value))

            _dataList = dataList.copy()

            for j in range(0, len(indexes)):
                for item in twoPart:
                    _column = columnList[item]
                    if j >= sourceCount:
                        _rx = indexes[j-sourceCount]+1
                        _dataList[item] = (
                            str(_ws.cell(row=_rx, column=_column).value))
                    else:
                        _rx = indexes[j]+1
                        _dataList[item] = (
                            str(ws.cell(row=_rx, column=_column).value))
                result.append(_dataList)
                _dataList = dataList.copy()

        #
        fullPath = exportSet(filePath, result, 'add',
                             str(parentIndex)+str(key))
        key += 1
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)


# 单纯累加
def readWbFromIndexAdd(wsList, eachFileDataCount, filePath, parentIndex):
    indexes = list(range(1, eachFileDataCount + 1))
    columnList = {'A': 2, 'B': 3, 'C': 4, 'D': 5}
    parts = ('A', 'B', 'C', 'D')
    #
    result = []
    for ws in wsList:
        #
        for j in range(0, len(indexes)):
            dataList = {'A': '', 'B': '', 'C': '', 'D': ''}
            for item in parts:
                _column = columnList[item]
                _rx = indexes[j]+1
                dataList[item] = (
                    str(ws.cell(row=_rx, column=_column).value))
            result.append(dataList)
        #
    fullPath = exportSet(filePath, result, 'leijia',
                         str(parentIndex))
    #
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fullPath)

    # FileFormat = 51 is for .xlsx extension
    wb.SaveAs(fullPath+"x", FileFormat=51)
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()
    #
    os.remove(fullPath)

# 批量excel set 区间处理


def excelSet():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    section = list(map(int, configuration['excelSection'].split(',')))
    excelCount = int(configuration['excelCount'])
    ui.textBrowser.append('计算数据中......')
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    # 排列组合 Cn2 后进行相加
    index = 1
    for path in cur:
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, excelCount+1)
        resultSet = readWbFromIndex(
            ws, indexes, "批量结果集"+str(index), section)
        file = open(filePath+"\\批量结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        index += 1
    # return result
    file.close()
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)

    root.destroy()
    root.mainloop()
    #
    pass


# 批量双区间
def TwosideExcelSet():
    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    leftSection = list(map(int, configuration['leftSection'].split(',')))
    rightSection = list(map(int, configuration['rightSection'].split(',')))
    excelCount = int(configuration['excelCount'])
    ui.textBrowser.append('批量双区间：计算数据中......')
    #
    root = Tk()
    cur = filedialog.askopenfilenames(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        return
    # 排列组合 Cn2 后进行相加
    index = 1
    for path in cur:
        # left
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, excelCount+1)
        resultSet = readWbFromIndex(
            ws, indexes, "批量左区间结果集"+str(index), leftSection)
        file = open(filePath+"\\批量左区间结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        # right
        wb = openpyxl.load_workbook(path)
        sheetsNames = wb.sheetnames
        ws = wb[sheetsNames[0]]
        indexes = range(1, excelCount+1)
        resultSet = readWbFromIndex(
            ws, indexes, "批量右区间结果集"+str(index), rightSection)
        file = open(filePath+"\\批量右区间结果集"+str(index) +
                    '_总计'+str(len(resultSet))+'.txt', 'w')
        file.write(','.join(list(resultSet)))
        file.flush()
        file.close()
        #
        index += 1
    ui.textBrowser.append('计算完成')
    os.startfile(filePath)

    root.destroy()
    root.mainloop()
    pass


# 原始数据随机组合
def originPaste():
    #
    root = Tk()
    cur = filedialog.askopenfilename(filetypes=[('xlsx files', '.xlsx')])
    if cur == '':
        ui.textBrowser.append('取消')
        return
    totalCount = int(enterbox("选择文件一共多少数据?", '确认', "0"))
    dataCount = int(enterbox("多少数据为一个新文件?", '确认', "0"))
    fileCount = int(enterbox("需要生成多少文件？", '确认', "0"))
    #
    filePath = diropenbox('结果存放目录')
    if filePath == None:
        ui.textBrowser.append('取消')
        root.destroy()
        return
        #
    #
    ui.textBrowser.append('开始计算....')
    wb = openpyxl.load_workbook(cur)
    sheetsNames = wb.sheetnames
    ws = wb[sheetsNames[0]]
    indexes = range(1, totalCount+1)
    aList = []
    bList = []
    cList = []
    dList = []
    # 获取excel abcd 4列总数据
    for i in range(0, len(indexes)):
        rx = indexes[i] + 1
        w1 = str(ws.cell(row=rx, column=1).value)
        # 千
        w2 = str(ws.cell(row=rx, column=2).value)
        # 百
        w3 = str(ws.cell(row=rx, column=3).value)
        # 十
        w4 = str(ws.cell(row=rx, column=4).value)
        # 个
        w5 = str(ws.cell(row=rx, column=5).value)
        if w2 == "None":
            break
        else:
            aList.append(w2)
            bList.append(w3)
            cList.append(w4)
            dList.append(w5)
    #
    for i in range(0, fileCount):
        #
        result = []
        for k in range(0, dataCount):
            indexList = random.sample(range(0, totalCount), 4)
            result.append(dict(A=aList[indexList[0]], B=bList[indexList[1]],
                               C=cList[indexList[2]], D=dList[indexList[3]]))
        fullPath = exportSet(filePath, result, 'zuhe',
                             str(i+1))
        #
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fullPath)

        # FileFormat = 51 is for .xlsx extension
        wb.SaveAs(fullPath+"x", FileFormat=51)
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
        #
        os.remove(fullPath)

    ui.textBrowser.append('计算完成')
    os.startfile(filePath)
    root.destroy()
    root.mainloop()
    #
    pass


def exportSet(path, data, bonus='', key=0):
    _workbook = xlwt.Workbook(encoding='utf-8')
    # 样式
    style = xlwt.XFStyle()  # 创建一个样式对象，初始化样式
    al = xlwt.Alignment()
    al.horz = 0x02      # 设置水平居中
    al.vert = 0x01      # 设置垂直居中
    style.alignment = al
    # 注意: 在add_sheet时, 置参数cell_overwrite_ok=True, 可以覆盖原单元格中数据。
    # cell_overwrite_ok默认为False, 覆盖的话, 会抛出异常.
    sheet = _workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
    # 获取并写入数据段信息
    sheet.write(0, 0, '姓名', style)
    sheet.write(0, 1, '数据1', style)
    sheet.write(0, 2, '数据2', style)
    sheet.write(0, 3, '数据3', style)
    sheet.write(0, 4, '数据4', style)
    for row in range(0, len(data)):
        for index in range(0, 4):
            sheet.write(row+1, 0, 'A'+str(row+1), style)
            sheet.write(row+1, index+1, int(data[row]
                                            [['A', 'B', 'C', 'D'][index]]), style)
    #
    if bonus == '':
        fullPath = path+'\\源文件'+'总计'+str(len(data))+'.xls'
    if bonus == 'add':
        fullPath = path+'\\叠加文件'+str(key)+'.xls'
    if bonus == 'leijia':
        fullPath = path+'\\指定数量累加文件'+str(key)+'.xls'
    if bonus == 'jiaocha':
        fullPath = path+'\\交叉文件'+str(key)+'.xls'
    if bonus == 'zuhe':
        fullPath = path+'\\组合文件'+str(key)+'.xls'
    _workbook.save(fullPath)
    return fullPath


if __name__ == '__main__':

    #
    passWord = passwordbox("请输入启动密码", '确认', "")
    nowDate = time.strftime("%Y%m%d", time.localtime())
    if passWord != nowDate+'lin':
        exit()

    with open("setting.json", "r") as f:
        setting = f.read()
    configuration = json.loads(setting)
    rightNumber = configuration['rightNumber']
    totalCount = int(configuration['totalCount'])
    hundredCount = int(configuration['hundredCount'])
    fiftyCount = int(configuration['fiftyCount'])
    handlerCount = int(configuration['handlerCount'])
    insideCount = int(configuration['insideCount'])
    isSingle = configuration['isSingle']
    isRange = configuration['isRange']
    qianRange = configuration['qianRange']
    baiRange = configuration['baiRange']
    shiRange = configuration['shiRange']
    geRange = configuration['geRange']
    setArray = configuration['setArray']
    #
    app = QApplication(sys.argv)
    MainWindow = QMainWindow()
    ui = window.Ui_MainWindow()
    ui.setupUi(MainWindow)
    #
    dialogWindow = QDialog()
    configui = config.Ui_config()
    configui.setupUi(dialogWindow)

    # 绑定dialog确认事件
    configui.buttonOk.accepted.connect(confirmConfig)
    configui.addSection.clicked.connect(addSection)
    configui.delSection.clicked.connect(delSection)
    MainWindow.show()
    # action绑定
    ui.actionjiaoji.triggered.connect(jiaoji)
    ui.actionchaji.triggered.connect(chaji)
    ui.actionbingji.triggered.connect(bingji)
    ui.actioncheck.triggered.connect(check)
    ui.actionBind.triggered.connect(mutilBind)
    ui.actionmutilCha.triggered.connect(mutilCha)
    # 随机求交
    ui.actiontxtRandom.triggered.connect(txtRandomTimes)
    ui.actiontxtRandomGroup.triggered.connect(txtRandomTimesGroup)
    # 随机求并
    ui.actionrandomPickBind.triggered.connect(txtRandomPickBind)
    ui.actionrandomGroupBind.triggered.connect(txtRandomGroupBind)
    # 随机抽
    ui.actiondiyPIckGroupInter.triggered.connect(txtDiyPickGroupInter)

    ui.clearButton.clicked.connect(clearTextBrowser)
    # 爬虫类绑定
    ui.actionloadingRecent.triggered.connect(loadRecent)
    ui.actionexportPast.triggered.connect(climb)
    # excel类绑定
    ui.actionconfig.triggered.connect(showConfig)
    ui.actionmainCheck.triggered.connect(mainCheck)
    ui.actionmini.triggered.connect(miniCheck)
    ui.actionSource.triggered.connect(sourceCheck)
    ui.actionAddSource.triggered.connect(addSourceCheck)
    ui.actionmutilAdd.triggered.connect(mutilAdd)
    ui.actionExcelSet.triggered.connect(excelSet)
    ui.actionactionTwoside.triggered.connect(TwosideExcelSet)
    ui.actionoriginPaste.triggered.connect(originPaste)
    sys.exit(app.exec_())
